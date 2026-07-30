#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=====================================================================================
 AI REPOSITORY AUDITOR  --  single standalone script
=====================================================================================

Performs a complete, zero-configuration, STATIC audit of any repository and emits:

    repository_summary.md      Markdown report (tables + summaries)
    repository_summary.html    Interactive dashboard (charts, search, filters,
                               sorting, dark mode, export buttons, responsive)
    repository_summary.json    Full machine-readable dump
    repository_summary.csv     Flat file inventory
    repository_summary.xlsx    Multi-sheet workbook (needs openpyxl; skipped if absent)

What it discovers automatically
-------------------------------
  * Every directory / file (recursive, incl. hidden dirs, minus the ignore list)
  * Language, category, purpose and a human-readable description per file
  * AI agents, agent types and purposes
  * AI providers (OpenAI, Azure, Anthropic, Gemini, Bedrock, Groq, Ollama, ...)
  * AI models + reference counts
  * AI SDKs / frameworks (LangChain, LangGraph, CrewAI, AutoGen, SK, LlamaIndex...)
  * Prompts (system/user/developer/inline/file/markdown/json/yaml/template)
  * API-key environment variables and where they are loaded from
  * Token configuration + static token/cost estimates
  * Request / rate-limit / retry / backoff / timeout configuration
  * Tools used by agents, and agentic workflow patterns
  * Import graph, dependency graph, unused / orphan files
  * Repository statistics (classes, functions, endpoints, tests, IaC, ...)

Guarantees
----------
  * STATIC ANALYSIS ONLY. Repository code is never imported or executed.
  * READ-ONLY. No file in the scanned repository is ever modified.
  * DETERMINISTIC. All collections are sorted; the same input yields the same output.
  * Anything that genuinely cannot be derived from source is reported verbatim as
    "Not Available from Source Code" instead of being guessed.

Usage
-----
    python ai_repo_audit.py                      # audit current directory
    python ai_repo_audit.py /path/to/repo
    python ai_repo_audit.py /path/to/repo -o ./audit_out
    python ai_repo_audit.py . --max-file-mb 5 --quiet
    python ai_repo_audit.py . --open             # open the dashboard in a browser
    python ai_repo_audit.py . --serve 8000       # view it at http://127.0.0.1:8000/

Viewing the dashboard
---------------------
repository_summary.html is a SELF-CONTAINED local file. Open it directly:

    file:///full/path/to/ai_audit_report/repository_summary.html

It is not a route inside your application. Requesting it from your own Flask/FastAPI/
Django server (e.g. http://localhost:5000/repository_summary.html) returns that
server's "Not Found" page, because the file lives on disk and not in that app's
routing table or static folder. Use --open, use --serve, or double-click the file.

Standard library only (openpyxl optional, only for the .xlsx report).
=====================================================================================
"""

from __future__ import annotations

import argparse
import csv
import datetime as _dt
import hashlib
import html as _html
import json
import os
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

VERSION = "1.0.0"
NA = "Not Available from Source Code"

# =====================================================================================
# SECTION 1 -- CONFIGURATION / KNOWLEDGE BASE
# =====================================================================================

IGNORE_DIRS: Set[str] = {
    ".git", ".terraform", "node_modules", "venv", "__pycache__", "dist", "build",
    "target", "coverage", ".cache", ".idea", ".vscode",
}
# Practical extras that are never source and would only add noise. Kept small and
# explicit so the "ignore only" contract above is honoured for real directories.
IGNORE_DIR_EXTRAS: Set[str] = {".venv", ".mypy_cache", ".pytest_cache", ".ruff_cache",
                               ".next", ".nuxt", ".svelte-kit", ".gradle", ".tox",
                               "site-packages", ".egg-info"}

BINARY_EXTS: Set[str] = {
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".ico", ".webp", ".svg", ".pdf", ".zip",
    ".gz", ".tar", ".tgz", ".bz2", ".xz", ".7z", ".rar", ".jar", ".war", ".ear",
    ".class", ".pyc", ".pyo", ".so", ".dll", ".dylib", ".exe", ".bin", ".o", ".a",
    ".woff", ".woff2", ".ttf", ".otf", ".eot", ".mp3", ".mp4", ".avi", ".mov", ".mkv",
    ".wav", ".flac", ".db", ".sqlite", ".sqlite3", ".parquet", ".pkl", ".pickle",
    ".h5", ".pt", ".pth", ".onnx", ".safetensors", ".bin", ".xlsx", ".xls", ".docx",
    ".pptx", ".lock",
}

EXT_LANGUAGE: Dict[str, str] = {
    ".py": "Python", ".pyi": "Python", ".ipynb": "Jupyter Notebook",
    ".java": "Java", ".kt": "Kotlin", ".kts": "Kotlin", ".scala": "Scala",
    ".sc": "Scala", ".groovy": "Groovy",
    ".cs": "C#", ".fs": "F#", ".vb": "VB.NET",
    ".js": "JavaScript", ".jsx": "JavaScript", ".mjs": "JavaScript",
    ".cjs": "JavaScript", ".ts": "TypeScript", ".tsx": "TypeScript",
    ".vue": "Vue", ".svelte": "Svelte",
    ".go": "Go", ".rs": "Rust", ".rb": "Ruby", ".php": "PHP", ".pl": "Perl",
    ".lua": "Lua", ".r": "R", ".jl": "Julia", ".dart": "Dart", ".swift": "Swift",
    ".m": "Objective-C", ".mm": "Objective-C++",
    ".c": "C", ".h": "C/C++ Header", ".cc": "C++", ".cpp": "C++", ".cxx": "C++",
    ".hpp": "C++ Header", ".hh": "C++ Header",
    ".sh": "Shell", ".bash": "Shell", ".zsh": "Shell", ".ksh": "Shell",
    ".ps1": "PowerShell", ".psm1": "PowerShell", ".psd1": "PowerShell",
    ".bat": "Batch", ".cmd": "Batch",
    ".yaml": "YAML", ".yml": "YAML", ".json": "JSON", ".json5": "JSON",
    ".jsonl": "JSON Lines", ".toml": "TOML", ".ini": "INI", ".cfg": "INI",
    ".conf": "Config", ".properties": "Properties", ".env": "Env",
    ".tf": "Terraform", ".tfvars": "Terraform", ".hcl": "HCL",
    ".bicep": "Bicep",
    ".md": "Markdown", ".markdown": "Markdown", ".mdx": "MDX", ".rst": "reStructuredText",
    ".txt": "Text", ".adoc": "AsciiDoc",
    ".xml": "XML", ".xsd": "XML", ".xsl": "XML", ".html": "HTML", ".htm": "HTML",
    ".css": "CSS", ".scss": "SCSS", ".sass": "SASS", ".less": "LESS",
    ".sql": "SQL", ".psql": "SQL", ".ddl": "SQL",
    ".graphql": "GraphQL", ".gql": "GraphQL", ".proto": "Protocol Buffers",
    ".jinja": "Jinja Template", ".jinja2": "Jinja Template", ".j2": "Jinja Template",
    ".hbs": "Handlebars", ".mustache": "Mustache", ".tmpl": "Template",
    ".tpl": "Template", ".prompt": "Prompt", ".prompty": "Prompt",
    ".gradle": "Gradle", ".make": "Makefile", ".mk": "Makefile",
    ".csproj": "MSBuild", ".sln": "Visual Studio Solution", ".fsproj": "MSBuild",
    ".gitignore": "Git Config", ".dockerignore": "Docker Config",
}

FILENAME_LANGUAGE: Dict[str, str] = {
    "dockerfile": "Dockerfile", "containerfile": "Dockerfile",
    "makefile": "Makefile", "gnumakefile": "Makefile",
    "jenkinsfile": "Jenkins Pipeline", "vagrantfile": "Ruby",
    "procfile": "Procfile", "gemfile": "Ruby", "rakefile": "Ruby",
    "cmakelists.txt": "CMake", "go.mod": "Go Modules", "go.sum": "Go Modules",
    "cargo.toml": "Rust Manifest", "package.json": "NPM Manifest",
    "requirements.txt": "Python Requirements", "pipfile": "Python Requirements",
    "pyproject.toml": "Python Project", "setup.py": "Python Setup",
    "setup.cfg": "Python Setup", "pom.xml": "Maven POM",
    ".env": "Env", ".env.example": "Env", ".env.sample": "Env", ".env.local": "Env",
    ".gitignore": "Git Config", ".dockerignore": "Docker Config",
    ".gitlab-ci.yml": "GitLab CI", "readme.md": "Markdown", "license": "Text",
}

SOURCE_LANGS: Set[str] = {
    "Python", "Java", "Kotlin", "Scala", "Groovy", "C#", "F#", "VB.NET", "JavaScript",
    "TypeScript", "Vue", "Svelte", "Go", "Rust", "Ruby", "PHP", "Perl", "Lua", "R",
    "Julia", "Dart", "Swift", "Objective-C", "Objective-C++", "C", "C++",
    "C/C++ Header", "C++ Header", "Shell", "PowerShell", "Batch", "SQL", "GraphQL",
    "Protocol Buffers", "Jupyter Notebook",
}
CONFIG_LANGS: Set[str] = {
    "YAML", "JSON", "JSON Lines", "TOML", "INI", "Config", "Properties", "Env", "XML",
    "NPM Manifest", "Python Requirements", "Python Project", "Python Setup",
    "Maven POM", "Gradle", "MSBuild", "Visual Studio Solution", "Go Modules",
    "Rust Manifest", "Git Config", "Docker Config", "CMake", "Procfile",
}
DOC_LANGS: Set[str] = {"Markdown", "MDX", "reStructuredText", "Text", "AsciiDoc"}
INFRA_LANGS: Set[str] = {
    "Terraform", "HCL", "Bicep", "Dockerfile", "Jenkins Pipeline", "GitLab CI",
    "Makefile",
}

# ---------------------------------------------------------------- providers ---------
# name -> (sdk hint, regex fragments, endpoint hints, env var hints)
PROVIDER_RULES: List[Dict[str, Any]] = [
    {"name": "Azure OpenAI", "sdk": "openai / azure-ai-openai",
     "patterns": [r"AzureOpenAI", r"AzureChatOpenAI", r"AzureOpenAIClient",
                  r"azure[_\-\.]openai", r"AZURE_OPENAI", r"openai\.azure\.com",
                  r"Azure\.AI\.OpenAI", r"api-version=\d{4}-\d{2}-\d{2}"],
     "env": [r"AZURE_OPENAI_\w+", r"AZURE_OPENAI_API_KEY", r"AZURE_OPENAI_KEY"]},
    {"name": "OpenAI", "sdk": "openai",
     "patterns": [r"\bfrom\s+openai\b", r"\bimport\s+openai\b", r"require\(['\"]openai['\"]\)",
                  r"\bOpenAI\(", r"ChatOpenAI", r"OpenAIClient", r"api\.openai\.com",
                  r"chat\.completions\.create", r"openai\.ChatCompletion"],
     "env": [r"OPENAI_API_KEY", r"OPENAI_ORG\w*", r"OPENAI_BASE_URL"]},
    {"name": "Anthropic", "sdk": "anthropic",
     "patterns": [r"\banthropic\b", r"Anthropic\(", r"ChatAnthropic", r"claude-",
                  r"api\.anthropic\.com", r"AnthropicBedrock"],
     "env": [r"ANTHROPIC_API_KEY", r"ANTHROPIC_\w+"]},
    {"name": "Google Gemini", "sdk": "google-generativeai / google-genai",
     "patterns": [r"google\.generativeai", r"google\.genai", r"GenerativeModel",
                  r"ChatGoogleGenerativeAI", r"generativelanguage\.googleapis\.com",
                  r"\bgemini[-_]", r"VertexAI", r"vertexai"],
     "env": [r"GEMINI_API_KEY", r"GOOGLE_API_KEY", r"GOOGLE_APPLICATION_CREDENTIALS",
             r"VERTEX_\w+"]},
    {"name": "AWS Bedrock", "sdk": "boto3 / aws-sdk bedrock",
     "patterns": [r"bedrock[-_]runtime", r"\bbedrock\b", r"BedrockChat", r"ChatBedrock",
                  r"invoke_model", r"BedrockRuntimeClient"],
     "env": [r"AWS_ACCESS_KEY_ID", r"AWS_SECRET_ACCESS_KEY", r"AWS_REGION",
             r"AWS_SESSION_TOKEN", r"BEDROCK_\w+"]},
    {"name": "Mistral", "sdk": "mistralai",
     "patterns": [r"mistralai", r"MistralClient", r"ChatMistralAI",
                  r"api\.mistral\.ai", r"\bmistral-", r"\bmixtral"],
     "env": [r"MISTRAL_API_KEY"]},
    {"name": "Groq", "sdk": "groq",
     "patterns": [r"\bfrom\s+groq\b", r"\bimport\s+groq\b", r"\bGroq\(", r"ChatGroq",
                  r"api\.groq\.com"],
     "env": [r"GROQ_API_KEY"]},
    {"name": "Cohere", "sdk": "cohere",
     "patterns": [r"\bcohere\b", r"ChatCohere", r"api\.cohere\.(ai|com)", r"command-r"],
     "env": [r"COHERE_API_KEY", r"CO_API_KEY"]},
    {"name": "Ollama", "sdk": "ollama",
     "patterns": [r"\bollama\b", r"ChatOllama", r"OllamaLLM", r"localhost:11434",
                  r"127\.0\.0\.1:11434"],
     "env": [r"OLLAMA_HOST", r"OLLAMA_\w+"]},
    {"name": "OpenRouter", "sdk": "openrouter / openai-compatible",
     "patterns": [r"openrouter", r"openrouter\.ai"],
     "env": [r"OPENROUTER_API_KEY"]},
    {"name": "HuggingFace", "sdk": "transformers / huggingface_hub",
     "patterns": [r"huggingface", r"transformers", r"AutoModel", r"AutoTokenizer",
                  r"pipeline\(\s*['\"]text-generation", r"InferenceClient",
                  r"HuggingFaceEndpoint", r"api-inference\.huggingface\.co"],
     "env": [r"HUGGINGFACE\w*", r"HF_TOKEN", r"HUGGINGFACEHUB_API_TOKEN"]},
    {"name": "DeepSeek", "sdk": "openai-compatible / deepseek",
     "patterns": [r"deepseek", r"api\.deepseek\.com"],
     "env": [r"DEEPSEEK_API_KEY"]},
    {"name": "Perplexity", "sdk": "openai-compatible / perplexity",
     "patterns": [r"perplexity", r"api\.perplexity\.ai", r"\bsonar[-_]"],
     "env": [r"PERPLEXITY_API_KEY", r"PPLX_API_KEY"]},
    {"name": "Meta Llama", "sdk": "llama-api / local weights",
     "patterns": [r"\bllama[-_]?\d", r"llama_cpp", r"llama-cpp", r"LlamaCpp",
                  r"meta-llama"],
     "env": [r"LLAMA_\w+", r"REPLICATE_API_TOKEN"]},
    {"name": "Local LLM", "sdk": "local runtime",
     "patterns": [r"llama_cpp", r"gpt4all", r"lmstudio", r"localai", r"vllm",
                  r"text-generation-webui", r"LocalAI"],
     "env": [r"LOCAL_LLM\w*", r"LLM_BASE_URL"]},
    {"name": "Custom API", "sdk": "custom http client",
     "patterns": [r"base_url\s*=\s*['\"]https?://(?!api\.openai|api\.anthropic)",
                  r"LLM_ENDPOINT", r"MODEL_ENDPOINT", r"INFERENCE_URL"],
     "env": [r"LLM_ENDPOINT", r"MODEL_ENDPOINT", r"INFERENCE_URL", r"CUSTOM_LLM\w*"]},
]

# ---------------------------------------------------------------- models ------------
MODEL_PATTERNS: List[Tuple[str, str]] = [
    # (regex, provider)
    (r"\bgpt-5(?:\.\d+)?(?:-[a-z0-9]+)*\b", "OpenAI"),
    (r"\bgpt-4\.1(?:-(?:mini|nano))?\b", "OpenAI"),
    (r"\bgpt-4o(?:-(?:mini|audio|realtime|search)[\w-]*)?\b", "OpenAI"),
    (r"\bgpt-4(?:-turbo|-32k|-vision-preview)?\b", "OpenAI"),
    (r"\bgpt-3\.5-turbo(?:-\w+)*\b", "OpenAI"),
    (r"\bo[134](?:-(?:mini|pro|preview))?\b", "OpenAI"),
    (r"\btext-embedding-(?:3-(?:small|large)|ada-002)\b", "OpenAI"),
    (r"\bwhisper-\d\b", "OpenAI"),
    (r"\bdall-e-\d\b", "OpenAI"),
    (r"\bclaude-(?:opus|sonnet|haiku)-[\w.\-]+\b", "Anthropic"),
    (r"\bclaude-\d(?:\.\d)?-(?:opus|sonnet|haiku)[\w.\-]*\b", "Anthropic"),
    (r"\bclaude-(?:instant|2|3)[\w.\-]*\b", "Anthropic"),
    (r"\bgemini-(?:\d(?:\.\d)?)-(?:pro|flash|ultra|nano)[\w.\-]*\b", "Google Gemini"),
    (r"\bgemini-(?:pro|ultra|flash)[\w.\-]*\b", "Google Gemini"),
    (r"\btext-bison[\w.\-]*\b", "Google Gemini"),
    (r"\bllama-?[234](?:\.\d)?[\w.\-]*\b", "Meta Llama"),
    (r"\bmeta-llama/[\w.\-]+\b", "Meta Llama"),
    (r"\bmistral-(?:tiny|small|medium|large|nemo)[\w.\-]*\b", "Mistral"),
    (r"\bmixtral-[\w.\-]+\b", "Mistral"),
    (r"\bcodestral[\w.\-]*\b", "Mistral"),
    (r"\bdeepseek-(?:chat|coder|reasoner|r1|v\d)[\w.\-]*\b", "DeepSeek"),
    (r"\bcommand-r(?:-plus)?\b", "Cohere"),
    (r"\bcommand(?:-light|-nightly)?\b", "Cohere"),
    (r"\bqwen[\w.\-]*\b", "Alibaba Qwen"),
    (r"\bphi-?[234][\w.\-]*\b", "Microsoft Phi"),
    (r"\bgemma-?\d?[\w.\-]*\b", "Google Gemma"),
    (r"\bgrok-[\w.\-]+\b", "xAI"),
    (r"\bsonar(?:-(?:pro|reasoning|deep-research))?\b", "Perplexity"),
    (r"\banthropic\.claude-[\w.\-:]+\b", "AWS Bedrock"),
    (r"\bamazon\.(?:titan|nova)-[\w.\-:]+\b", "AWS Bedrock"),
    (r"\bcohere\.command[\w.\-:]*\b", "AWS Bedrock"),
    (r"\bnomic-embed-text\b", "Ollama"),
]

# rough public pricing (USD per 1M tokens) -- used ONLY for a clearly-labelled estimate
MODEL_PRICING: Dict[str, Tuple[float, float]] = {
    "gpt-5": (1.25, 10.0), "gpt-4.1": (2.0, 8.0), "gpt-4.1-mini": (0.4, 1.6),
    "gpt-4.1-nano": (0.1, 0.4), "gpt-4o": (2.5, 10.0), "gpt-4o-mini": (0.15, 0.6),
    "gpt-4-turbo": (10.0, 30.0), "gpt-4": (30.0, 60.0), "gpt-3.5-turbo": (0.5, 1.5),
    "o3": (2.0, 8.0), "o3-mini": (1.1, 4.4), "o4-mini": (1.1, 4.4), "o1": (15.0, 60.0),
    "claude-opus": (15.0, 75.0), "claude-sonnet": (3.0, 15.0), "claude-haiku": (0.8, 4.0),
    "gemini-1.5-pro": (1.25, 5.0), "gemini-1.5-flash": (0.075, 0.3),
    "gemini-2.0-flash": (0.1, 0.4), "gemini-2.5-pro": (1.25, 10.0),
    "mistral-large": (2.0, 6.0), "mistral-small": (0.2, 0.6),
    "deepseek-chat": (0.27, 1.1), "command-r-plus": (2.5, 10.0), "command-r": (0.15, 0.6),
}

MODEL_CONTEXT: Dict[str, int] = {
    "gpt-5": 400000, "gpt-4.1": 1047576, "gpt-4.1-mini": 1047576, "gpt-4.1-nano": 1047576,
    "gpt-4o": 128000, "gpt-4o-mini": 128000, "gpt-4-turbo": 128000, "gpt-4": 8192,
    "gpt-3.5-turbo": 16385, "o3": 200000, "o3-mini": 200000, "o4-mini": 200000,
    "o1": 200000, "claude-opus": 200000, "claude-sonnet": 200000, "claude-haiku": 200000,
    "gemini-1.5-pro": 2000000, "gemini-1.5-flash": 1000000, "gemini-2.0-flash": 1048576,
    "gemini-2.5-pro": 1048576, "mistral-large": 128000, "mixtral": 32768,
    "llama-3": 8192, "deepseek-chat": 64000, "command-r-plus": 128000,
}

# ---------------------------------------------------------------- SDKs --------------
SDK_RULES: List[Tuple[str, List[str]]] = [
    ("OpenAI SDK", [r"\bopenai\b", r"from\s+openai", r"OpenAI\(", r"AzureOpenAI\("]),
    ("Anthropic SDK", [r"\banthropic\b", r"Anthropic\("]),
    ("Google GenAI SDK", [r"google\.generativeai", r"google\.genai", r"GenerativeModel"]),
    ("AWS Bedrock SDK", [r"bedrock-runtime", r"BedrockRuntime", r"bedrock_runtime"]),
    ("LangChain", [r"\blangchain\b", r"from\s+langchain", r"langchain[_\.]core",
                   r"langchain[_\.]community", r"@langchain/"]),
    ("LangGraph", [r"\blanggraph\b", r"StateGraph", r"MessageGraph", r"langgraph\."]),
    ("Semantic Kernel", [r"semantic_kernel", r"Microsoft\.SemanticKernel",
                         r"semantic-kernel", r"\bKernel\.CreateBuilder"]),
    ("AutoGen", [r"\bautogen\b", r"AssistantAgent", r"UserProxyAgent",
                 r"GroupChatManager", r"autogen_agentchat"]),
    ("CrewAI", [r"\bcrewai\b", r"from\s+crewai", r"\bCrew\(", r"\bTask\(.*agent="]),
    ("LlamaIndex", [r"llama_index", r"llamaindex", r"VectorStoreIndex",
                    r"ServiceContext", r"@llamaindex/"]),
    ("Haystack", [r"\bhaystack\b", r"haystack\.components", r"Pipeline\(\)"]),
    ("DSPy", [r"\bdspy\b", r"dspy\.(Module|Predict|ChainOfThought|Signature)"]),
    ("Transformers", [r"\btransformers\b", r"AutoModelFor", r"AutoTokenizer",
                      r"pipeline\("]),
    ("LiteLLM", [r"\blitellm\b", r"litellm\.completion"]),
    ("Instructor", [r"\binstructor\b", r"instructor\.(patch|from_openai)"]),
    ("Ollama SDK", [r"\bollama\b", r"ollama\.(chat|generate)"]),
    ("OpenRouter SDK", [r"openrouter"]),
    ("Vercel AI SDK", [r"\bai/rsc\b", r"from\s+['\"]ai['\"]", r"@ai-sdk/",
                       r"streamText\(", r"generateText\("]),
    ("Pydantic AI", [r"pydantic_ai", r"pydantic-ai"]),
    ("Guidance", [r"\bguidance\b", r"guidance\.(gen|select)"]),
    ("Sentence Transformers", [r"sentence_transformers", r"SentenceTransformer"]),
    ("Spring AI", [r"org\.springframework\.ai", r"spring-ai"]),
    ("LangChain4j", [r"dev\.langchain4j", r"langchain4j"]),
]

# ---------------------------------------------------------------- agents ------------
AGENT_SIGNALS: List[Tuple[str, str]] = [
    (r"class\s+(\w*Agent\w*)\b", "Agent Class"),
    (r"class\s+(\w*Assistant\w*)\b", "Assistant Class"),
    (r"class\s+(\w*Orchestrator\w*)\b", "Orchestrator Class"),
    (r"class\s+(\w*Copilot\w*)\b", "Copilot Class"),
    (r"class\s+(\w*Bot\w*)\b", "Bot Class"),
    (r"class\s+(\w*Chain\w*)\b", "Chain Class"),
    (r"class\s+(\w*Planner\w*)\b", "Planner Class"),
    (r"class\s+(\w*Worker\w*)\b", "Worker Class"),
    (r"class\s+(\w*Crew\w*)\b", "Crew Class"),
    (r"class\s+(\w*Supervisor\w*)\b", "Supervisor Class"),
    (r"(?:def|async\s+def|function|func|public\s+\w+)\s+(\w*agent\w*)\s*\(", "Agent Function"),
    (r"(?:def|async\s+def|function|func)\s+(run_\w*agent\w*)\s*\(", "Agent Runner"),
]
AGENT_CONSTRUCTORS: List[Tuple[str, str]] = [
    (r"\bAgent\s*\(\s*(?:name\s*=\s*)?['\"]([\w \-]+)['\"]", "SDK Agent"),
    (r"\bAssistantAgent\s*\(\s*(?:name\s*=\s*)?['\"]([\w \-]+)['\"]", "AutoGen Assistant Agent"),
    (r"\bUserProxyAgent\s*\(\s*(?:name\s*=\s*)?['\"]([\w \-]+)['\"]", "AutoGen User Proxy Agent"),
    (r"\bConversableAgent\s*\(\s*(?:name\s*=\s*)?['\"]([\w \-]+)['\"]", "AutoGen Conversable Agent"),
    (r"\bcreate_react_agent\s*\(", "LangGraph ReAct Agent"),
    (r"\bcreate_openai_functions_agent\s*\(", "LangChain Functions Agent"),
    (r"\bcreate_tool_calling_agent\s*\(", "LangChain Tool-Calling Agent"),
    (r"\bAgentExecutor\s*\(", "LangChain Agent Executor"),
    (r"\binitialize_agent\s*\(", "LangChain Agent"),
    (r"\bStateGraph\s*\(", "LangGraph State Machine Agent"),
    (r"\bcrewai\.Agent\s*\(", "CrewAI Agent"),
    (r"\bKernel\s*\(\)", "Semantic Kernel Agent"),
    (r"\bChatCompletionAgent\s*\(", "Semantic Kernel Chat Agent"),
    (r"\bSwarm\s*\(", "Swarm Agent"),
]
AI_CALL_PATTERNS: List[str] = [
    r"chat\.completions\.create", r"completions\.create", r"ChatCompletion\.create",
    r"messages\.create", r"generate_content", r"invoke_model", r"\.invoke\(",
    r"\.ainvoke\(", r"\.stream\(", r"\.astream\(", r"\.predict\(", r"\.run\(",
    r"litellm\.completion", r"ollama\.chat", r"generateText\(", r"streamText\(",
    r"GetChatMessageContentAsync", r"CompleteChatAsync", r"createChatCompletion",
    r"\.chat\(", r"\.complete\(", r"embeddings\.create", r"embed_query",
]

# ---------------------------------------------------------------- tools -------------
TOOL_RULES: List[Tuple[str, List[str]]] = [
    ("Web Search", [r"tavily", r"serpapi", r"duckduckgo", r"google_search", r"bing_search",
                    r"web_search", r"SerperDev", r"brave_search"]),
    ("Browser", [r"playwright", r"selenium", r"puppeteer", r"browser_tool", r"BrowserBase"]),
    ("Vector Store", [r"pinecone", r"weaviate", r"qdrant", r"chromadb", r"\bchroma\b",
                      r"milvus", r"faiss", r"pgvector", r"lancedb", r"opensearch",
                      r"AzureSearch", r"azure\.search"]),
    ("RAG", [r"\brag\b", r"retriev(?:er|al)", r"VectorStoreRetriever", r"as_retriever",
             r"RetrievalQA", r"context_documents"]),
    ("Database", [r"psycopg2?", r"sqlalchemy", r"pymongo", r"mysql", r"postgres",
                  r"jdbc:", r"EntityFramework", r"gorm", r"sqlite3", r"redis"]),
    ("Calculator", [r"calculator", r"\bnumexpr\b", r"math_tool", r"llm_math"]),
    ("Email", [r"smtplib", r"sendgrid", r"\bmailgun\b", r"javax\.mail", r"nodemailer"]),
    ("Slack", [r"slack_sdk", r"\bslack\b", r"WebClient\(.*slack", r"@slack/"]),
    ("GitHub", [r"\bPyGithub\b", r"github3", r"octokit", r"api\.github\.com", r"\bgh api\b"]),
    ("Azure", [r"azure\.\w+", r"Azure\.\w+", r"az\s+\w+", r"azure-identity",
               r"DefaultAzureCredential"]),
    ("AWS", [r"\bboto3\b", r"aws-sdk", r"AmazonS3", r"\bs3\b", r"lambda_client"]),
    ("GCP", [r"google\.cloud", r"gcloud", r"bigquery"]),
    ("Filesystem", [r"FileManagementToolkit", r"read_file", r"write_file", r"os\.walk",
                    r"pathlib", r"fs\.readFile"]),
    ("Shell", [r"subprocess", r"os\.system", r"ShellTool", r"child_process", r"exec\("]),
    ("Python REPL", [r"PythonREPL", r"python_repl", r"exec\(", r"PythonAstREPLTool"]),
    ("Memory", [r"ConversationBufferMemory", r"ConversationSummaryMemory", r"\bmemory\b",
                r"MemorySaver", r"chat_history", r"checkpointer"]),
    ("HTTP / API", [r"\brequests\.\w+", r"httpx", r"aiohttp", r"axios", r"fetch\(",
                    r"RestTemplate", r"HttpClient"]),
    ("Queue", [r"\bcelery\b", r"rabbitmq", r"\bkafka\b", r"sqs", r"servicebus", r"\bbullmq\b"]),
    ("Scheduler", [r"\bcron\b", r"APScheduler", r"schedule\.every", r"Quartz"]),
]

WORKFLOW_RULES: List[Tuple[str, List[str]]] = [
    ("Planning", [r"\bplan(?:ner|ning)?\b", r"create_plan", r"decompose", r"task_list",
                  r"SequentialPlanner", r"plan_and_execute"]),
    ("Execution", [r"\bexecut(?:e|or|ion)\b", r"AgentExecutor", r"run_step", r"act\("]),
    ("Reflection", [r"\breflect(?:ion)?\b", r"self_critique", r"critic", r"reviewer",
                    r"self_refine"]),
    ("Retry", [r"\bretry\b", r"@retry", r"tenacity", r"max_retries", r"Polly", r"retryPolicy"]),
    ("Backoff", [r"backoff", r"exponential_backoff", r"wait_exponential", r"jitter"]),
    ("Evaluation", [r"\beval(?:uate|uation)\b", r"ragas", r"deepeval", r"scorer", r"\bjudge\b"]),
    ("Memory", [r"ConversationBufferMemory", r"MemorySaver", r"chat_history",
                r"vector_memory", r"checkpointer", r"\bmemory\b"]),
    ("RAG", [r"retriev", r"RetrievalQA", r"as_retriever", r"rerank", r"\brag\b"]),
    ("Tool Calling", [r"tool_calls", r"tool_choice", r"bind_tools", r"@tool\b",
                      r"ToolNode", r"StructuredTool"]),
    ("Function Calling", [r"function_call", r"functions\s*=\s*\[", r"tools\s*=\s*\[",
                          r"FunctionDefinition", r"parallel_tool_calls"]),
    ("Streaming", [r"stream\s*=\s*True", r"\.astream", r"streamText", r"SSE",
                   r"text/event-stream", r"yield\s+chunk"]),
    ("Multi-agent", [r"GroupChat", r"multi_agent", r"agents\s*=\s*\[", r"\bCrew\(",
                     r"handoff", r"Swarm\("]),
    ("Supervisor", [r"supervisor", r"router_agent", r"orchestrator"]),
    ("Coordinator", [r"coordinator", r"dispatcher", r"delegat"]),
    ("Guardrails", [r"guardrail", r"moderation", r"content_filter", r"nemoguardrails"]),
    ("Human in the Loop", [r"human_in_the_loop", r"interrupt\(", r"approval",
                           r"HumanApproval"]),
]

# ---------------------------------------------------------------- env keys ----------
ENV_KEY_RE = re.compile(
    r"\b([A-Z][A-Z0-9]*(?:_[A-Z0-9]+)*_(?:API_)?(?:KEY|TOKEN|SECRET|PASSWORD|CREDENTIALS|"
    r"ENDPOINT|URL|URI|REGION|DEPLOYMENT|VERSION|MODEL|HOST|ID))\b")
AI_ENV_HINT = re.compile(
    r"(OPENAI|AZURE|ANTHROPIC|CLAUDE|GEMINI|GOOGLE|VERTEX|BEDROCK|AWS|GROQ|COHERE|"
    r"MISTRAL|OLLAMA|OPENROUTER|HUGGINGFACE|HF_|DEEPSEEK|PERPLEXITY|PPLX|LLM|MODEL|"
    r"REPLICATE|TOGETHER|FIREWORKS|XAI|GROK|LANGCHAIN|LANGSMITH|LANGFUSE|TAVILY|"
    r"PINECONE|WEAVIATE|QDRANT)")

ENV_LOADERS: List[Tuple[str, str]] = [
    (r"load_dotenv|dotenv\.config|DotEnv", ".env file (dotenv loader)"),
    (r"os\.environ|os\.getenv|process\.env|System\.getenv|Environment\.GetEnvironmentVariable|"
     r"std::env::var|os\.Getenv", "Process environment"),
    (r"KeyVault|SecretClient|azure\.keyvault", "Azure Key Vault"),
    (r"SecretsManager|secretsmanager|ssm\.get_parameter", "AWS Secrets Manager / SSM"),
    (r"pydantic_settings|BaseSettings", "Pydantic Settings"),
    (r"ConfigMap|kind:\s*Secret", "Kubernetes ConfigMap / Secret"),
    (r"variable\s+\"|TF_VAR_", "Terraform variables"),
]

# ---------------------------------------------------------------- purposes ----------
# ordered rules: (purpose, path/name regex, content regex)
PURPOSE_RULES: List[Tuple[str, Optional[str], Optional[str]]] = [
    ("AI Agent", r"(agent|copilot|assistant|crew|swarm|orchestrat)", None),
    ("AI Agent", None, r"class\s+\w*Agent\w*|AgentExecutor|AssistantAgent|crewai|StateGraph"),
    ("Prompt", r"(prompt|template.*prompt|system_message|persona)", None),
    ("Prompt", None, r"SYSTEM_PROMPT|system_prompt|PromptTemplate|ChatPromptTemplate"),
    ("LLM", r"(llm|model|completion|inference|chat)", None),
    ("LLM", None, r"chat\.completions\.create|generate_content|invoke_model|messages\.create"),
    ("Memory", r"(memory|history|conversation|context_store|checkpoint)", None),
    ("Tool", r"(tool|plugin|function_?call|skill)", None),
    ("Workflow", r"(workflow|pipeline|graph|chain|dag|flow|saga)", None),
    ("RAG / Vector Store", r"(rag|retriev|embedding|vector|index|ingest|chunk)", None),
    ("Testing", r"(^|/)(tests?|spec|__tests__)(/|$)|(_test|test_|\.test\.|\.spec\.)", None),
    ("Docker", r"(dockerfile|docker-compose|containerfile|\.dockerignore)", None),
    ("Kubernetes", r"(k8s|kubernetes|helm|chart|manifests?)", r"apiVersion:|kind:\s*(Deployment|Service|Pod|ConfigMap|Ingress)"),
    ("Kubernetes", r"(k8s|kubernetes|helm|chart)", None),
    ("Terraform", r"\.tf$|\.tfvars$|terraform", None),
    ("CI/CD Pipeline", r"(\.github/workflows|\.gitlab-ci|azure-pipelines|jenkinsfile|circleci|buildkite|\.drone)", None),
    ("Deployment", r"(deploy|release|rollout|provision)", None),
    ("Infrastructure", r"(infra|infrastructure|bicep|cloudformation|pulumi|ansible)", None),
    ("API", r"(api|rest|graphql|endpoint|resource)", None),
    ("Controller", r"(controller|handler|route|router|view)", None),
    ("API", None, r"@app\.(get|post|put|delete|patch)|@router\.|@RestController|app\.(get|post)\(|FastAPI\(|express\(\)"),
    ("Database", r"(model|entity|schema|migration|repositor|dao|orm|db)", None),
    ("Database", None, r"CREATE TABLE|sqlalchemy|SELECT .* FROM|@Entity|mongoose\.Schema"),
    ("Service", r"(service|manager|provider|usecase|business|domain)", None),
    ("Authentication", r"(auth|login|signin|oauth|jwt|token_service|identity|session)", None),
    ("Authorization", r"(rbac|permission|policy|acl|authoriz|guard)", None),
    ("Logging", r"(log|logger|logging|audit)", None),
    ("Monitoring", r"(monitor|health|probe|trace|otel|telemetry|observab)", None),
    ("Metrics", r"(metric|prometheus|grafana|stats|usage)", None),
    ("Scheduler", r"(schedul|cron|job|timer|task_runner|worker)", None),
    ("Queue", r"(queue|broker|kafka|rabbit|sqs|pubsub|topic|consumer|producer)", None),
    ("CLI", r"(cli|command|console|main|__main__|entrypoint|bin/)", None),
    ("CLI", None, r"argparse|click\.command|cobra\.Command|commander|typer\.Typer"),
    ("Frontend", r"(component|page|ui|view|frontend|client|web/|static|styles?)", None),
    ("Frontend", None, r"import\s+React|useState\(|<template>|@Component\("),
    ("Backend", r"(server|backend|app|main|host|startup|program)", None),
    ("Configuration", r"(config|settings|conf|options|env|properties|\.ini|\.toml)", None),
    ("Documentation", r"(readme|docs?/|changelog|contributing|license|architecture|adr)", None),
    ("Utility", r"(util|helper|common|shared|lib|tools?/|misc)", None),
    ("Package Manifest", r"(package\.json|requirements|pyproject|pom\.xml|build\.gradle|cargo\.toml|go\.mod|\.csproj)", None),
]

DESCRIPTION_HINTS: Dict[str, str] = {
    "AI Agent": "Implements or wires an AI agent: builds the LLM client, holds the agent loop and coordinates tools.",
    "Prompt": "Holds prompt text/templates used to instruct the language model.",
    "LLM": "Performs direct language-model / inference calls and handles model responses.",
    "Memory": "Stores or retrieves conversation state, history or agent memory.",
    "Tool": "Defines callable tools/functions that an agent can invoke.",
    "Workflow": "Defines orchestration flow: steps, graph nodes, chains or pipeline stages.",
    "RAG / Vector Store": "Handles embeddings, chunking, indexing and retrieval for grounded generation.",
    "Testing": "Automated test code validating behaviour of the codebase.",
    "Docker": "Container build/compose definition for packaging and running the service.",
    "Kubernetes": "Kubernetes manifest describing workloads, services or configuration.",
    "Terraform": "Terraform IaC defining cloud resources and their configuration.",
    "CI/CD Pipeline": "Continuous integration/delivery pipeline definition.",
    "Deployment": "Deployment scripts or descriptors for shipping the application.",
    "Infrastructure": "Infrastructure-as-code / platform provisioning assets.",
    "API": "Exposes HTTP/GraphQL endpoints and request-response contracts.",
    "Controller": "Routes inbound requests to services and shapes responses.",
    "Database": "Data models, schema, migrations or persistence access code.",
    "Service": "Business/domain logic invoked by controllers and agents.",
    "Authentication": "Verifies caller identity (login, tokens, credentials).",
    "Authorization": "Enforces access control, roles, scopes and policies.",
    "Logging": "Log configuration and structured logging helpers.",
    "Monitoring": "Health checks, tracing and observability wiring.",
    "Metrics": "Collects and exposes counters, usage and performance metrics.",
    "Scheduler": "Timed or recurring job execution.",
    "Queue": "Asynchronous messaging: producers, consumers and brokers.",
    "CLI": "Command-line entrypoint and argument handling.",
    "Frontend": "User-interface code, components, views or styling.",
    "Backend": "Application bootstrap / server composition root.",
    "Configuration": "Configuration values, settings and environment wiring.",
    "Documentation": "Human-readable documentation for the repository.",
    "Utility": "Reusable helper functions shared across modules.",
    "Package Manifest": "Declares dependencies and build metadata for the package.",
    "General": "General-purpose source or asset file.",
}

TEST_PATH_RE = re.compile(r"(^|/)(tests?|spec|specs|__tests__|testing)(/|$)|"
                          r"(^|/)(test_[^/]+|[^/]+_test|[^/]+\.test|[^/]+\.spec)\.\w+$", re.I)

AI_KEYWORD_RE = re.compile(
    r"(openai|anthropic|claude|gemini|bedrock|langchain|langgraph|llama|mistral|groq|"
    r"cohere|ollama|huggingface|transformers|deepseek|perplexity|semantic_kernel|"
    r"semantickernel|autogen|crewai|llama_index|llamaindex|haystack|litellm|dspy|"
    r"instructor|openrouter|\bllm\b|\bgpt-|prompt|embedding|completion|inference|"
    r"chat_model|chatmodel|agent|vertexai|azure_openai|azureopenai)", re.I)

# =====================================================================================
# SECTION 2 -- DATA MODEL
# =====================================================================================


@dataclass
class FileRecord:
    serial: int = 0
    rel_path: str = ""
    name: str = ""
    ext: str = ""
    language: str = "Unknown"
    category: str = "Other"
    purpose: str = "General"
    description: str = ""
    size_bytes: int = 0
    size_human: str = ""
    lines_total: int = 0
    lines_code: int = 0
    is_used: str = NA
    is_referenced: str = "No"
    is_ai_related: str = "No"
    dependencies: List[str] = field(default_factory=list)
    internal_dependencies: List[str] = field(default_factory=list)
    classes: List[str] = field(default_factory=list)
    functions: List[str] = field(default_factory=list)
    endpoints: List[str] = field(default_factory=list)
    providers: List[str] = field(default_factory=list)
    models: List[str] = field(default_factory=list)
    sdks: List[str] = field(default_factory=list)
    tools: List[str] = field(default_factory=list)
    workflows: List[str] = field(default_factory=list)
    env_vars: List[str] = field(default_factory=list)
    agents: List[str] = field(default_factory=list)
    prompt_count: int = 0
    ai_calls: int = 0
    module_names: List[str] = field(default_factory=list)
    sha1: str = ""
    binary: bool = False


# =====================================================================================
# SECTION 3 -- HELPERS
# =====================================================================================


def human_size(n: int) -> str:
    step = 1024.0
    val = float(n)
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if val < step or unit == "TB":
            return f"{val:.0f} {unit}" if unit == "B" else f"{val:.1f} {unit}"
        val /= step
    return f"{n} B"


def uniq(seq: Iterable[str]) -> List[str]:
    return sorted({s for s in seq if s})


def safe_read(path: str, max_bytes: int) -> Tuple[str, bool]:
    """Return (text, is_binary). Never raises."""
    try:
        if os.path.getsize(path) > max_bytes:
            with open(path, "rb") as fh:
                raw = fh.read(max_bytes)
        else:
            with open(path, "rb") as fh:
                raw = fh.read()
    except OSError:
        return "", True
    if b"\x00" in raw[:8000]:
        return "", True
    for enc in ("utf-8", "utf-16", "latin-1"):
        try:
            return raw.decode(enc), False
        except (UnicodeDecodeError, LookupError):
            continue
    return "", True


def strip_noise(text: str) -> str:
    """Remove comment-only noise for code-line counting (best effort, language agnostic)."""
    out = []
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        if s.startswith(("#", "//", "/*", "*", "--", "<!--", ";")):
            continue
        out.append(line)
    return "\n".join(out)


def findall(patterns: Sequence[str], text: str, flags: int = re.I) -> List[str]:
    hits: List[str] = []
    for p in patterns:
        try:
            for m in re.finditer(p, text, flags):
                hits.append(m.group(0))
        except re.error:
            continue
    return hits


def count_matches(patterns: Sequence[str], text: str, flags: int = re.I) -> int:
    total = 0
    for p in patterns:
        try:
            total += len(re.findall(p, text, flags))
        except re.error:
            pass
    return total


def estimate_tokens(text: str) -> int:
    """~4 characters per token heuristic (documented as an estimate everywhere)."""
    return max(0, int(len(text) / 4))


def norm_model(model: str) -> str:
    return model.strip().strip("\"'").lower()


def price_for(model: str) -> Optional[Tuple[float, float]]:
    m = norm_model(model)
    best: Optional[Tuple[str, Tuple[float, float]]] = None
    for key, val in MODEL_PRICING.items():
        if key in m and (best is None or len(key) > len(best[0])):
            best = (key, val)
    return best[1] if best else None


def context_for(model: str) -> Optional[int]:
    m = norm_model(model)
    best: Optional[Tuple[str, int]] = None
    for key, val in MODEL_CONTEXT.items():
        if key in m and (best is None or len(key) > len(best[0])):
            best = (key, val)
    return best[1] if best else None


# =====================================================================================
# SECTION 4 -- LANGUAGE / CATEGORY / PURPOSE CLASSIFICATION
# =====================================================================================


def detect_language(name: str, ext: str) -> str:
    low = name.lower()
    if low in FILENAME_LANGUAGE:
        return FILENAME_LANGUAGE[low]
    for key, lang in FILENAME_LANGUAGE.items():
        if low == key or low.startswith(key + "."):
            return lang
    if low.startswith("dockerfile"):
        return "Dockerfile"
    if low.startswith(".env"):
        return "Env"
    if ext in EXT_LANGUAGE:
        return EXT_LANGUAGE[ext]
    if not ext and low.isupper():
        return "Text"
    return "Unknown"


def detect_category(rel: str, language: str) -> str:
    if TEST_PATH_RE.search(rel):
        return "Test"
    if language in INFRA_LANGS:
        return "Infrastructure"
    if language in DOC_LANGS:
        return "Documentation"
    if language in CONFIG_LANGS:
        return "Configuration"
    if language in SOURCE_LANGS:
        return "Source"
    if language in ("Jinja Template", "Handlebars", "Mustache", "Template", "Prompt"):
        return "Template"
    if language in ("HTML", "CSS", "SCSS", "SASS", "LESS", "MDX"):
        return "Frontend Asset"
    return "Other"


def detect_purpose(rel: str, name: str, text: str, category: str) -> str:
    path_key = rel.lower()
    body = text[:200000]
    if category == "Test":
        return "Testing"
    for purpose, path_re, content_re in PURPOSE_RULES:
        if path_re and re.search(path_re, path_key, re.I):
            return purpose
        if content_re and body and re.search(content_re, body):
            return purpose
    if category == "Documentation":
        return "Documentation"
    if category == "Configuration":
        return "Configuration"
    return "General"


def build_description(rec: FileRecord, text: str) -> str:
    """Prefer a real docstring/header comment; otherwise synthesise from signals."""
    doc = extract_doc_summary(text, rec.language)
    if doc:
        return doc
    parts: List[str] = [DESCRIPTION_HINTS.get(rec.purpose, DESCRIPTION_HINTS["General"])]
    extra: List[str] = []
    if rec.agents:
        extra.append("Defines agent(s): " + ", ".join(rec.agents[:4]))
    if rec.providers:
        extra.append("Provider(s): " + ", ".join(rec.providers))
    if rec.models:
        extra.append("Model(s): " + ", ".join(rec.models[:4]))
    if rec.sdks:
        extra.append("SDK(s): " + ", ".join(rec.sdks[:4]))
    if rec.endpoints:
        extra.append(f"{len(rec.endpoints)} HTTP endpoint(s)")
    if rec.classes or rec.functions:
        extra.append(f"{len(rec.classes)} class(es), {len(rec.functions)} function(s)")
    if extra:
        parts.append(" | ".join(extra) + ".")
    return " ".join(parts)


def extract_doc_summary(text: str, language: str) -> str:
    if not text:
        return ""
    head = text[:4000]
    m = re.search(r'^\s*(?:#!.*\n)?(?:#.*\n)*\s*(?:"""|\'\'\')(.{10,400}?)(?:"""|\'\'\')',
                  head, re.S)
    if m:
        return " ".join(m.group(1).split())[:300]
    m = re.search(r"/\*\*(.{10,400}?)\*/", head, re.S)
    if m:
        cleaned = re.sub(r"^\s*\*\s?", "", m.group(1), flags=re.M)
        return " ".join(cleaned.split())[:300]
    if language in DOC_LANGS:
        for line in head.splitlines():
            s = line.strip().lstrip("#").strip()
            if len(s) > 15 and not s.startswith(("!", "[", "|", "-", "=")):
                return s[:300]
    lines = head.splitlines()[:6]
    comments = [l.strip().lstrip("#/;-* ").strip() for l in lines
                if l.strip().startswith(("#", "//", ";", "--"))]
    comments = [c for c in comments if len(c) > 15 and "!/" not in c]
    if comments:
        return " ".join(comments)[:300]
    return ""


# =====================================================================================
# SECTION 5 -- STRUCTURAL EXTRACTION (classes, functions, imports, endpoints)
# =====================================================================================

CLASS_RES = [
    r"^\s*class\s+([A-Za-z_]\w*)",                                   # py/js/ts/php/scala
    r"\b(?:public|private|internal|protected|abstract|sealed|final|static)?\s*"
    r"(?:class|interface|record|enum|struct)\s+([A-Za-z_]\w*)",      # java/c#/kotlin
    r"^\s*type\s+([A-Za-z_]\w*)\s+struct",                           # go
    r"^\s*(?:pub\s+)?struct\s+([A-Za-z_]\w*)",                       # rust
    r"^\s*(?:pub\s+)?trait\s+([A-Za-z_]\w*)",                        # rust
]
FUNC_RES = [
    r"^\s*(?:async\s+)?def\s+([A-Za-z_]\w*)",                        # python
    r"^\s*(?:export\s+)?(?:async\s+)?function\s+([A-Za-z_]\w*)",     # js/ts
    r"^\s*(?:export\s+)?const\s+([A-Za-z_]\w*)\s*=\s*(?:async\s*)?\(", # js arrow
    r"^\s*func\s+(?:\([^)]*\)\s*)?([A-Za-z_]\w*)",                   # go
    r"^\s*(?:pub\s+)?(?:async\s+)?fn\s+([A-Za-z_]\w*)",              # rust
    r"^\s*(?:public|private|protected|internal|static|final|override|suspend|virtual|async)"
    r"[\w<>\[\],\s\.]*\s+([A-Za-z_]\w*)\s*\([^;]*\)\s*\{",           # java/c#/kotlin
    r"^\s*(?:fun)\s+([A-Za-z_]\w*)",                                 # kotlin
    r"^\s*(?:function)\s+([A-Za-z_]\w*)\s*\{",                       # shell
    r"^\s*([A-Za-z_]\w*)\s*\(\)\s*\{",                               # shell
]
IMPORT_RES = [
    r"^\s*import\s+([\w\.]+)",
    r"^\s*from\s+([\w\.]+)\s+import",
    r"^\s*using\s+([\w\.]+)\s*;",
    r"require\(\s*['\"]([^'\"]+)['\"]\s*\)",
    r"^\s*import\s+.*?from\s+['\"]([^'\"]+)['\"]",
    r"^\s*import\s+['\"]([^'\"]+)['\"]",
    r"^\s*use\s+([\w:]+)",
    r"^\s*#include\s+[<\"]([^>\"]+)[>\"]",
    r"^\s*source\s+([\w\./\-]+)",
    r"^\s*@?Import\(\s*['\"]?([\w\./\-]+)",
]
ENDPOINT_RES = [
    r"@(?:app|router|api|bp|blueprint)\.(get|post|put|delete|patch|head|options)\(\s*['\"]([^'\"]+)['\"]",
    r"@(?:Get|Post|Put|Delete|Patch)Mapping\(\s*(?:value\s*=\s*)?['\"]([^'\"]+)['\"]",
    r"@RequestMapping\(\s*(?:value\s*=\s*)?['\"]([^'\"]+)['\"]",
    r"\[Http(Get|Post|Put|Delete|Patch)\(\s*\"([^\"]*)\"\s*\)\]",
    r"(?:app|router)\.(get|post|put|delete|patch)\(\s*['\"]([^'\"]+)['\"]",
    r"http\.HandleFunc\(\s*\"([^\"]+)\"",
    r"\.route\(\s*['\"]([^'\"]+)['\"]",
    r"path\(\s*['\"]([^'\"]+)['\"]",
]


def extract_structure(text: str) -> Tuple[List[str], List[str], List[str], List[str]]:
    classes: List[str] = []
    functions: List[str] = []
    imports: List[str] = []
    endpoints: List[str] = []
    body = text[:400000]
    for pat in CLASS_RES:
        classes += re.findall(pat, body, re.M)
    for pat in FUNC_RES:
        functions += re.findall(pat, body, re.M)
    for pat in IMPORT_RES:
        imports += re.findall(pat, body, re.M)
    for pat in ENDPOINT_RES:
        for m in re.finditer(pat, body, re.M):
            groups = [g for g in m.groups() if g]
            if not groups:
                continue
            if len(groups) >= 2:
                endpoints.append(f"{groups[0].upper()} {groups[1]}")
            else:
                endpoints.append(groups[0])
    keywords = {"if", "for", "while", "switch", "catch", "return", "new", "class",
                "function", "def", "get", "set", "try", "else", "do", "main_"}
    functions = [f for f in functions if f.lower() not in keywords]
    return uniq(classes), uniq(functions), uniq(imports), uniq(endpoints)


# =====================================================================================
# SECTION 6 -- AI DETECTION
# =====================================================================================


def detect_providers(text: str) -> List[Tuple[str, str]]:
    found: List[Tuple[str, str]] = []
    for rule in PROVIDER_RULES:
        if count_matches(rule["patterns"], text) > 0:
            found.append((rule["name"], rule["sdk"]))
    names = {n for n, _ in found}
    # Azure OpenAI implies OpenAI SDK usage, keep both only when both signals are strong
    if "Azure OpenAI" in names and "OpenAI" in names:
        if count_matches([r"api\.openai\.com", r"OpenAI\(\s*api_key"], text) == 0:
            found = [f for f in found if f[0] != "OpenAI"]
    return sorted(set(found))


def detect_models(text: str) -> List[Tuple[str, str, int]]:
    """Match model ids, keeping only the longest match at each text position so that
    e.g. 'gpt-4.1-mini' is not also reported as 'gpt-4'."""
    spans: List[Tuple[int, int, str, str]] = []
    for pat, provider in MODEL_PATTERNS:
        for m in re.finditer(pat, text, re.I):
            spans.append((m.start(), m.end(), m.group(0), provider))
    spans.sort(key=lambda s: (s[0], -(s[1] - s[0]), s[2]))
    kept: List[Tuple[int, int, str, str]] = []
    for s in spans:
        if any(s[0] >= k[0] and s[1] <= k[1] for k in kept):
            continue  # fully contained in a longer match
        kept.append(s)
    out: Dict[str, Tuple[str, int]] = {}
    for _, _, name, provider in kept:
        key = name.lower()
        prev = out.get(key)
        out[key] = (prev[0] if prev else provider, (prev[1] if prev else 0) + 1)
    return sorted([(k, v[0], v[1]) for k, v in out.items()])


def detect_sdks(text: str) -> List[str]:
    return sorted({name for name, pats in SDK_RULES if count_matches(pats, text) > 0})


def detect_tools(text: str) -> List[str]:
    return sorted({name for name, pats in TOOL_RULES if count_matches(pats, text) > 0})


def detect_workflows(text: str) -> List[str]:
    return sorted({name for name, pats in WORKFLOW_RULES if count_matches(pats, text) > 0})


def detect_env_vars(text: str) -> List[str]:
    hits = set(ENV_KEY_RE.findall(text))
    for pat in (r"os\.getenv\(\s*['\"]([A-Za-z_][\w]*)['\"]",
                r"os\.environ(?:\.get)?[\(\[]\s*['\"]([A-Za-z_][\w]*)['\"]",
                r"process\.env\.([A-Za-z_]\w*)",
                r"process\.env\[['\"]([A-Za-z_]\w*)['\"]\]",
                r"System\.getenv\(\s*\"([A-Za-z_]\w*)\"",
                r"Environment\.GetEnvironmentVariable\(\s*\"([A-Za-z_]\w*)\"",
                r"os\.Getenv\(\s*\"([A-Za-z_]\w*)\"",
                r"env::var\(\s*\"([A-Za-z_]\w*)\"",
                r"^\s*([A-Z][A-Z0-9_]{2,})\s*=", ):
        hits.update(re.findall(pat, text, re.M))
    noise = re.compile(r"(PROMPT|TEMPLATE|MESSAGE|INSTRUCTION|PERSONA|SCHEMA|QUERY|SQL)$")
    return sorted({h for h in hits if len(h) > 2 and h.isupper() and not noise.search(h)})


def detect_agents(text: str, rel_path: str) -> List[Dict[str, str]]:
    agents: List[Dict[str, str]] = []
    seen: Set[str] = set()
    for pat, kind in AGENT_SIGNALS:
        for m in re.finditer(pat, text, re.M):
            name = m.group(1)
            if not name or name.lower() in {"agent", "baseagent"} and kind == "Agent Function":
                pass
            key = name.lower()
            if key in seen:
                continue
            seen.add(key)
            agents.append({"name": name, "type": kind})
    for pat, kind in AGENT_CONSTRUCTORS:
        for m in re.finditer(pat, text):
            name = m.group(1) if m.groups() else ""
            if not name:
                name = os.path.splitext(os.path.basename(rel_path))[0]
            key = name.lower() + "|" + kind
            if key in seen:
                continue
            seen.add(key)
            agents.append({"name": name, "type": kind})
    return agents


def detect_prompts(text: str, rel: str, language: str) -> List[Dict[str, Any]]:
    """Locate prompts: dedicated prompt files + inline prompt strings/templates."""
    prompts: List[Dict[str, Any]] = []
    low_rel = rel.lower()
    is_prompt_file = bool(re.search(r"(prompt|persona|instruction|system_message|template)",
                                    low_rel))

    def add(name: str, kind: str, content: str, line: int) -> None:
        content = content.strip()
        if len(content) < 12:
            return
        prompts.append({
            "name": name, "type": kind, "line": line,
            "chars": len(content), "estimated_tokens": estimate_tokens(content),
            "preview": " ".join(content.split())[:160],
        })

    # 1) whole-file prompts
    if is_prompt_file and language in (DOC_LANGS | {"Prompt", "Jinja Template", "Template",
                                                    "Handlebars", "Mustache", "Text"}):
        kind = {"Markdown": "Markdown Prompt File"}.get(language, "Prompt File")
        add(os.path.basename(rel), kind, text, 1)
        return prompts

    # 2) structured prompts inside json/yaml
    if language in ("JSON", "YAML") and is_prompt_file:
        for m in re.finditer(r"^[ \t\-]*['\"]?([\w .\-]*(?:prompt|instruction|system|persona|"
                             r"message|template)[\w .\-]*)['\"]?\s*[:=]\s*(.+)$",
                             text, re.I | re.M):
            add(m.group(1).strip(), f"{language} Prompt", m.group(2), text[:m.start()].count("\n") + 1)
        if prompts:
            return prompts

    # 3) named constants holding prompts
    for m in re.finditer(
            r"([A-Za-z_][\w\.]*(?:PROMPT|prompt|INSTRUCTION|instruction|SYSTEM_MESSAGE|"
            r"TEMPLATE|template|PERSONA|persona)[\w]*)\s*[:=]\s*"
            r"(?:f|r|rb|b)?(\"\"\"|'''|`|\"|')(.*?)\2", text, re.S):
        name, _, body = m.group(1), m.group(2), m.group(3)
        kind = "System Prompt" if re.search(r"system", name, re.I) else \
               "User Prompt" if re.search(r"user", name, re.I) else \
               "Developer Prompt" if re.search(r"developer", name, re.I) else \
               "Prompt Template" if re.search(r"template", name, re.I) else "Named Prompt"
        add(name, kind, body, text[:m.start()].count("\n") + 1)

    # 4) role-tagged inline messages
    for m in re.finditer(r"['\"]role['\"]\s*:\s*['\"](system|user|assistant|developer)['\"]"
                         r"\s*,\s*['\"]content['\"]\s*:\s*"
                         r"(?:f|r)?(\"\"\"|'''|`|\"|')(.*?)\2", text, re.S | re.I):
        role = m.group(1).capitalize()
        add(f"{role} message", f"{role} Prompt", m.group(3), text[:m.start()].count("\n") + 1)

    # 5) framework prompt builders
    for m in re.finditer(r"(?:SystemMessage|HumanMessage|AIMessage|SystemMessagePromptTemplate|"
                         r"HumanMessagePromptTemplate|PromptTemplate\.from_template|"
                         r"ChatPromptTemplate\.from_template|from_messages)\s*\(\s*"
                         r"(?:content\s*=\s*)?(\"\"\"|'''|`|\"|')(.*?)\1", text, re.S):
        add("Framework prompt", "Prompt Template", m.group(2),
            text[:m.start()].count("\n") + 1)

    # 6) prompt-ish long strings in obvious prompt files
    if is_prompt_file and not prompts:
        for m in re.finditer(r"(\"\"\"|''')(.*?)\1", text, re.S):
            add("Inline prompt", "Inline Prompt", m.group(2),
                text[:m.start()].count("\n") + 1)

    # dedupe deterministically
    out, seen = [], set()
    for p in prompts:
        key = (p["name"], p["preview"][:60])
        if key in seen:
            continue
        seen.add(key)
        out.append(p)
    return out


NUM_PARAM_RES: Dict[str, List[str]] = {
    "max_tokens": [r"max_tokens\s*[:=]\s*(\d+)", r"maxTokens\s*[:=]\s*(\d+)",
                   r"max_output_tokens\s*[:=]\s*(\d+)", r"MaxTokens\s*=\s*(\d+)",
                   r"max_completion_tokens\s*[:=]\s*(\d+)", r"maxOutputTokens\s*[:=]\s*(\d+)"],
    "temperature": [r"temperature\s*[:=]\s*([0-9.]+)", r"Temperature\s*=\s*([0-9.]+)"],
    "top_p": [r"top_p\s*[:=]\s*([0-9.]+)", r"topP\s*[:=]\s*([0-9.]+)", r"TopP\s*=\s*([0-9.]+)"],
    "timeout": [r"timeout\s*[:=]\s*([0-9.]+)", r"Timeout\s*=\s*([0-9.]+)",
                r"request_timeout\s*[:=]\s*([0-9.]+)"],
    "max_retries": [r"max_retries\s*[:=]\s*(\d+)", r"maxRetries\s*[:=]\s*(\d+)",
                    r"retries\s*[:=]\s*(\d+)", r"retry_count\s*[:=]\s*(\d+)"],
    "context_window": [r"context_window\s*[:=]\s*(\d+)", r"num_ctx\s*[:=]\s*(\d+)",
                       r"max_context\w*\s*[:=]\s*(\d+)"],
}
RATE_LIMIT_RES: Dict[str, List[str]] = {
    "requests_per_minute": [r"(?:rpm|requests_per_minute|requestsPerMinute|"
                            r"rate_limit_rpm)\s*[:=]\s*(\d+)"],
    "requests_per_day": [r"(?:rpd|requests_per_day|requestsPerDay|daily_limit)\s*[:=]\s*(\d+)"],
    "requests_per_month": [r"(?:requests_per_month|monthly_limit|monthly_quota)\s*[:=]\s*(\d+)"],
    "tokens_per_minute": [r"(?:tpm|tokens_per_minute|tokensPerMinute)\s*[:=]\s*(\d+)"],
    "concurrency": [r"(?:max_concurrency|concurrency|max_workers|parallelism)\s*[:=]\s*(\d+)"],
}


def extract_numeric_params(text: str) -> Dict[str, List[str]]:
    out: Dict[str, List[str]] = {}
    for key, pats in {**NUM_PARAM_RES, **RATE_LIMIT_RES}.items():
        vals: List[str] = []
        for p in pats:
            vals += re.findall(p, text)
        if vals:
            out[key] = sorted(set(vals), key=lambda v: (len(v), v))
    return out


# =====================================================================================
# SECTION 7 -- SCANNER
# =====================================================================================


class RepositoryAuditor:
    def __init__(self, root: str, max_file_mb: float = 3.0, quiet: bool = False,
                 follow_symlinks: bool = False):
        self.root = os.path.abspath(root)
        self.max_bytes = int(max_file_mb * 1024 * 1024)
        self.quiet = quiet
        self.follow_symlinks = follow_symlinks
        self.files: List[FileRecord] = []
        self.dirs: List[str] = []
        self.errors: List[str] = []
        self.agents: List[Dict[str, Any]] = []
        self.prompts: List[Dict[str, Any]] = []
        self.providers: Dict[str, Dict[str, Any]] = {}
        self.models: Dict[str, Dict[str, Any]] = {}
        self.sdks: Dict[str, List[str]] = defaultdict(list)
        self.tools: Dict[str, List[str]] = defaultdict(list)
        self.workflows: Dict[str, List[str]] = defaultdict(list)
        self.env_keys: Dict[str, Dict[str, Any]] = {}
        self.params: Dict[str, Dict[str, List[str]]] = {}
        self.runtime_metrics: List[Dict[str, str]] = []
        self.scan_started = _dt.datetime.now()

    # ---------------------------------------------------------------- logging
    def log(self, msg: str) -> None:
        if not self.quiet:
            print(msg, flush=True)

    # ---------------------------------------------------------------- walk
    def walk(self) -> None:
        ignored = IGNORE_DIRS | IGNORE_DIR_EXTRAS
        for dirpath, dirnames, filenames in os.walk(self.root, followlinks=self.follow_symlinks):
            dirnames[:] = sorted(d for d in dirnames
                                 if d not in ignored and not d.endswith(".egg-info"))
            filenames.sort()
            rel_dir = os.path.relpath(dirpath, self.root)
            if rel_dir != ".":
                self.dirs.append(rel_dir.replace(os.sep, "/"))
            for fn in filenames:
                full = os.path.join(dirpath, fn)
                if os.path.islink(full) and not self.follow_symlinks:
                    continue
                try:
                    if not os.path.isfile(full):
                        continue
                except OSError:
                    continue
                self.analyze_file(full)
        self.files.sort(key=lambda r: r.rel_path)
        for i, rec in enumerate(self.files, 1):
            rec.serial = i

    # ---------------------------------------------------------------- per file
    def analyze_file(self, full: str) -> None:
        rel = os.path.relpath(full, self.root).replace(os.sep, "/")
        name = os.path.basename(full)
        ext = os.path.splitext(name)[1].lower()
        rec = FileRecord(rel_path=rel, name=name, ext=ext or "(none)")
        try:
            rec.size_bytes = os.path.getsize(full)
        except OSError as exc:
            self.errors.append(f"{rel}: {exc}")
            return
        rec.size_human = human_size(rec.size_bytes)
        rec.language = detect_language(name, ext)
        rec.category = detect_category(rel, rec.language)

        if ext in BINARY_EXTS:
            rec.binary = True
            rec.description = "Binary or non-text asset (not statically analysed)."
            rec.purpose = detect_purpose(rel, name, "", rec.category)
            rec.is_used = NA
            self.files.append(rec)
            return

        text, is_bin = safe_read(full, self.max_bytes)
        if is_bin:
            rec.binary = True
            rec.description = "Binary or undecodable file (not statically analysed)."
            rec.purpose = detect_purpose(rel, name, "", rec.category)
            self.files.append(rec)
            return

        rec.sha1 = hashlib.sha1(text.encode("utf-8", "ignore")).hexdigest()[:12]
        rec.lines_total = text.count("\n") + (1 if text and not text.endswith("\n") else 0)
        rec.lines_code = len(strip_noise(text).splitlines())

        classes, functions, imports, endpoints = extract_structure(text)
        rec.classes, rec.functions, rec.endpoints = classes, functions, endpoints
        rec.dependencies = imports[:200]

        providers = detect_providers(text)
        rec.providers = [p for p, _ in providers]
        models = detect_models(text)
        rec.models = [m for m, _, _ in models]
        rec.sdks = detect_sdks(text)
        rec.tools = detect_tools(text)
        rec.workflows = detect_workflows(text)
        rec.env_vars = detect_env_vars(text)
        rec.ai_calls = count_matches(AI_CALL_PATTERNS, text, 0)

        prompts = detect_prompts(text, rel, rec.language)
        rec.prompt_count = len(prompts)

        ai_score = (len(rec.providers) * 3 + len(rec.models) * 3 + len(rec.sdks) * 3 +
                    min(rec.ai_calls, 5) * 2 + rec.prompt_count)
        keyword_hits = len(AI_KEYWORD_RE.findall(text[:150000]))
        is_ai = ai_score >= 3 or (keyword_hits >= 4 and (rec.providers or rec.sdks or rec.models))
        rec.is_ai_related = "Yes" if is_ai else "No"

        detected_agents: List[Dict[str, str]] = []
        if rec.category != "Test" and (
                is_ai or rec.ai_calls or
                re.search(r"agent|assistant|orchestrat|crew|copilot", rel, re.I)):
            detected_agents = detect_agents(text, rel)
            # only keep agents in files with genuine AI signal
            if not (rec.providers or rec.sdks or rec.models or rec.ai_calls):
                detected_agents = [a for a in detected_agents
                                   if re.search(r"agent|assistant|crew|copilot|orchestrat",
                                                a["name"], re.I)]
            detected_agents = self._dedupe_agents(detected_agents, rel)
        rec.agents = [a["name"] for a in detected_agents]

        rec.purpose = detect_purpose(rel, name, text, rec.category)
        if is_ai and rec.purpose in ("General", "Service", "Backend", "Utility"):
            if detected_agents:
                rec.purpose = "AI Agent"
            elif rec.prompt_count:
                rec.purpose = "Prompt"
            elif rec.ai_calls or rec.models:
                rec.purpose = "LLM"
        rec.description = build_description(rec, text)

        params = extract_numeric_params(text)
        if params:
            self.params[rel] = params

        self._register_ai(rel, rec, providers, models, prompts, detected_agents, text)
        self._register_runtime_metrics(rel, text)
        self.files.append(rec)

    # ---------------------------------------------------------------- registries
    def _register_ai(self, rel: str, rec: FileRecord,
                     providers: List[Tuple[str, str]],
                     models: List[Tuple[str, str, int]],
                     prompts: List[Dict[str, Any]],
                     agents: List[Dict[str, str]], text: str) -> None:
        for pname, sdk in providers:
            entry = self.providers.setdefault(pname, {
                "provider": pname, "sdk": sdk, "endpoints": set(), "api_versions": set(),
                "auth_methods": set(), "env_vars": set(), "files": set(),
            })
            entry["files"].add(rel)
            for url in re.findall(r"https?://[\w\.\-]+(?:/[\w\.\-/{}$]*)?", text):
                if re.search(r"(openai|anthropic|google|gemini|bedrock|amazonaws|groq|"
                             r"cohere|mistral|ollama|11434|openrouter|huggingface|deepseek|"
                             r"perplexity|azure)", url, re.I):
                    entry["endpoints"].add(url.rstrip("\"',"))
            for v in re.findall(r"api[_\-]?version\s*[:=]\s*['\"]?([\w\-\.]+)", text, re.I):
                entry["api_versions"].add(v)
            for v in re.findall(r"['\"](\d{4}-\d{2}-\d{2}(?:-preview)?)['\"]", text):
                entry["api_versions"].add(v)
            for pat, label in ((r"api[_\-]?key", "API Key"),
                               (r"Bearer\s", "Bearer Token"),
                               (r"DefaultAzureCredential|ManagedIdentity|AzureCliCredential",
                                "Azure Managed Identity / AAD"),
                               (r"boto3|AWS_ACCESS_KEY|sigv4|Signature", "AWS IAM / SigV4"),
                               (r"service_account|GOOGLE_APPLICATION_CREDENTIALS",
                                "GCP Service Account"),
                               (r"no auth|localhost:11434", "None / Local")):
                if re.search(pat, text, re.I):
                    entry["auth_methods"].add(label)
            rule = next((r for r in PROVIDER_RULES if r["name"] == pname), None)
            if rule:
                for epat in rule["env"]:
                    for ev in re.findall(epat, text):
                        entry["env_vars"].add(ev if isinstance(ev, str) else ev[0])

        for mname, mprovider, cnt in models:
            entry = self.models.setdefault(mname, {
                "model": mname, "provider": mprovider, "files": set(), "references": 0,
            })
            entry["files"].add(rel)
            entry["references"] += cnt

        for sdk in rec.sdks:
            self.sdks[sdk].append(rel)
        for tool in rec.tools:
            self.tools[tool].append(rel)
        for wf in rec.workflows:
            self.workflows[wf].append(rel)

        for ev in rec.env_vars:
            provider = self._provider_for_env(ev)
            entry = self.env_keys.setdefault(ev, {
                "variable": ev, "provider": provider, "files": set(), "loaded_from": set(),
                "ai_related": bool(AI_ENV_HINT.search(ev)),
            })
            entry["files"].add(rel)
            if entry["provider"] == "Unknown" and provider != "Unknown":
                entry["provider"] = provider
            for pat, label in ENV_LOADERS:
                if re.search(pat, text):
                    entry["loaded_from"].add(label)
            if rel.lower().endswith((".env", ".env.example", ".env.sample")) or \
                    os.path.basename(rel).startswith(".env"):
                entry["loaded_from"].add(f"{os.path.basename(rel)} file")

        for p in prompts:
            self.prompts.append({**p, "file": rel, "purpose": self._prompt_purpose(p),
                                 "agent": ", ".join(rec.agents) if rec.agents else NA})

        for a in agents:
            self.agents.append({
                "name": a["name"], "type": a["type"], "file": rel,
                "providers": rec.providers, "models": rec.models, "sdks": rec.sdks,
                "tools": rec.tools, "workflows": rec.workflows,
                "prompts": [p["name"] for p in prompts],
                "env_vars": [e for e in rec.env_vars if AI_ENV_HINT.search(e)],
                "ai_calls": rec.ai_calls,
                "purpose": self._agent_purpose(a, rec),
            })

    @staticmethod
    def _dedupe_agents(agents: List[Dict[str, str]], rel: str) -> List[Dict[str, str]]:
        """Collapse framework-constructor hits that merely restate a class already found
        in the same file (e.g. StateGraph() inside supervisor.py -> SupervisorAgent)."""
        explicit = [a for a in agents if "Class" in a["type"] or "Function" in a["type"]]
        stem = os.path.splitext(os.path.basename(rel))[0].lower().replace("_", "")
        out: List[Dict[str, str]] = []
        for a in agents:
            if a in explicit:
                out.append(a)
                continue
            nm = a["name"].lower().replace("_", "")
            if explicit and (nm == stem or any(nm in e["name"].lower() or
                                               e["name"].lower().startswith(nm)
                                               for e in explicit)):
                continue  # framework construction of an already-named agent
            out.append(a)
        return out

    @staticmethod
    def _prompt_purpose(p: Dict[str, Any]) -> str:
        t = p["type"].lower()
        if "system" in t:
            return "Defines the model's role, rules and behaviour"
        if "user" in t:
            return "User-turn message sent to the model"
        if "developer" in t:
            return "Developer instruction layer"
        if "template" in t:
            return "Parameterised prompt rendered at runtime"
        if "file" in t:
            return "Standalone prompt asset loaded by the application"
        return "Instruction text supplied to the language model"

    @staticmethod
    def _agent_purpose(a: Dict[str, str], rec: FileRecord) -> str:
        n = a["name"].lower()
        bits: List[str] = []
        if "supervisor" in n or "orchestr" in n or "coordinat" in n or "router" in n:
            bits.append("Coordinates and routes work across other agents")
        elif "plan" in n:
            bits.append("Decomposes goals into an executable plan")
        elif "research" in n or "search" in n:
            bits.append("Gathers and synthesises external information")
        elif "review" in n or "critic" in n or "eval" in n:
            bits.append("Reviews / critiques output for quality")
        elif "write" in n or "summar" in n or "report" in n:
            bits.append("Generates written content from inputs")
        elif "code" in n or "dev" in n:
            bits.append("Produces or modifies code")
        elif "chat" in n or "support" in n or "assistant" in n:
            bits.append("Conversational assistant handling user turns")
        else:
            bits.append("Executes an LLM-backed task loop")
        if rec.tools:
            bits.append("tools: " + ", ".join(rec.tools[:4]))
        if rec.workflows:
            bits.append("patterns: " + ", ".join(rec.workflows[:4]))
        return "; ".join(bits)

    @staticmethod
    def _provider_for_env(ev: str) -> str:
        table = [
            ("AZURE_OPENAI", "Azure OpenAI"), ("AZURE", "Azure"), ("OPENAI", "OpenAI"),
            ("ANTHROPIC", "Anthropic"), ("CLAUDE", "Anthropic"),
            ("GEMINI", "Google Gemini"), ("VERTEX", "Google Gemini"),
            ("GOOGLE_API", "Google Gemini"), ("GOOGLE_APPLICATION", "Google Cloud"),
            ("BEDROCK", "AWS Bedrock"), ("AWS", "AWS"), ("GROQ", "Groq"),
            ("COHERE", "Cohere"), ("CO_API", "Cohere"), ("MISTRAL", "Mistral"),
            ("OLLAMA", "Ollama"), ("OPENROUTER", "OpenRouter"),
            ("HUGGINGFACE", "HuggingFace"), ("HF_", "HuggingFace"),
            ("DEEPSEEK", "DeepSeek"), ("PERPLEXITY", "Perplexity"), ("PPLX", "Perplexity"),
            ("REPLICATE", "Replicate"), ("TOGETHER", "Together AI"),
            ("FIREWORKS", "Fireworks AI"), ("XAI", "xAI"), ("GROK", "xAI"),
            ("LANGCHAIN", "LangChain / LangSmith"), ("LANGSMITH", "LangSmith"),
            ("LANGFUSE", "Langfuse"), ("TAVILY", "Tavily (tool)"),
            ("PINECONE", "Pinecone (vector store)"), ("QDRANT", "Qdrant (vector store)"),
            ("WEAVIATE", "Weaviate (vector store)"),
        ]
        up = ev.upper()
        for key, prov in table:
            if key in up:
                return prov
        return "Unknown"

    def _register_runtime_metrics(self, rel: str, text: str) -> None:
        """Look for evidence that the repo records real usage at runtime."""
        signals = [
            (r"usage\.(?:total_tokens|prompt_tokens|completion_tokens)", "Token usage read from API response"),
            (r"(?:prompt|completion|total)_tokens", "Token counters present in code"),
            (r"tiktoken|count_tokens|token_counter", "Token counting implemented"),
            (r"x-ratelimit-(?:remaining|limit)", "Rate-limit headers inspected"),
            (r"prometheus|Counter\(|Histogram\(|opentelemetry", "Metrics exporter present"),
            (r"langsmith|langfuse|helicone|wandb|mlflow", "LLM observability platform wired"),
            (r"cost_?(?:per_?token|estimate|usd)", "Cost tracking implemented"),
        ]
        for pat, label in signals:
            if re.search(pat, text, re.I):
                self.runtime_metrics.append({"file": rel, "signal": label})

    # ---------------------------------------------------------------- graphs
    def build_graphs(self) -> Dict[str, Any]:
        """Resolve imports to internal files and compute references / orphans."""
        by_module: Dict[str, List[str]] = defaultdict(list)
        by_stem: Dict[str, List[str]] = defaultdict(list)
        for rec in self.files:
            stem = os.path.splitext(rec.name)[0]
            by_stem[stem.lower()].append(rec.rel_path)
            dotted = os.path.splitext(rec.rel_path)[0].replace("/", ".")
            by_module[dotted.lower()].append(rec.rel_path)
            parts = dotted.split(".")
            for i in range(len(parts)):
                by_module[".".join(parts[i:]).lower()].append(rec.rel_path)
            rec.module_names = uniq([stem, dotted])

        referenced: Set[str] = set()
        edges: List[Tuple[str, str]] = []
        for rec in self.files:
            resolved: Set[str] = set()
            for dep in rec.dependencies:
                d = dep.strip().strip("./").replace("/", ".").replace("\\", ".")
                d = re.sub(r"\.(js|ts|jsx|tsx|py|mjs|cjs)$", "", d, flags=re.I)
                cands = by_module.get(d.lower()) or by_stem.get(d.split(".")[-1].lower())
                if not cands:
                    continue
                for c in cands:
                    if c != rec.rel_path:
                        resolved.add(c)
            # textual references to filenames (covers configs, docker, k8s, docs)
            rec.internal_dependencies = sorted(resolved)
            for target in resolved:
                referenced.add(target)
                edges.append((rec.rel_path, target))

        # second pass: literal path/name mentions inside any text file
        name_index: Dict[str, str] = {}
        for rec in self.files:
            name_index.setdefault(rec.name, rec.rel_path)
        for rec in self.files:
            if rec.binary:
                continue
            full = os.path.join(self.root, rec.rel_path)
            text, is_bin = safe_read(full, min(self.max_bytes, 400000))
            if is_bin or not text:
                continue
            for other in self.files:
                if other.rel_path == rec.rel_path:
                    continue
                if other.rel_path in referenced:
                    continue
                if other.name and len(other.name) > 3 and other.name in text:
                    referenced.add(other.rel_path)
                    edges.append((rec.rel_path, other.rel_path))

        entrypoint_re = re.compile(
            r"(^|/)(main|index|app|server|program|__main__|manage|cli|setup|conftest)\.\w+$|"
            r"(^|/)(readme|license|changelog|contributing)|"
            r"(dockerfile|makefile|jenkinsfile|\.github/|\.gitlab-ci|requirements|"
            r"package\.json|pyproject|go\.mod|cargo\.toml|pom\.xml|\.tf$|\.env)", re.I)

        orphans: List[str] = []
        for rec in self.files:
            is_entry = bool(entrypoint_re.search(rec.rel_path))
            is_test = rec.category == "Test"
            is_doc = rec.category in ("Documentation", "Configuration", "Infrastructure")
            rec.is_referenced = "Yes" if rec.rel_path in referenced else "No"
            if rec.rel_path in referenced:
                rec.is_used = "Yes"
            elif is_entry:
                rec.is_used = "Yes (entrypoint / conventional file)"
            elif is_test:
                rec.is_used = "Yes (test executed by test runner)"
            elif is_doc:
                rec.is_used = "Likely (declarative asset)"
            else:
                rec.is_used = "No reference found"
                orphans.append(rec.rel_path)

        agent_graph = []
        for a in self.agents:
            agent_graph.append({
                "agent": a["name"], "file": a["file"],
                "providers": a["providers"] or [NA],
                "models": a["models"] or [NA],
                "sdks": a["sdks"] or [NA],
                "tools": a["tools"] or [],
            })
        model_graph = [{"model": m, "provider": d["provider"],
                        "files": sorted(d["files"])} for m, d in sorted(self.models.items())]

        return {
            "import_edges": sorted(set(edges)),
            "file_dependency_graph": {r.rel_path: r.internal_dependencies
                                      for r in self.files if r.internal_dependencies},
            "agent_dependency_graph": agent_graph,
            "model_dependency_graph": model_graph,
            "orphan_files": sorted(orphans),
            "unused_files": sorted(orphans),
            "dead_code_candidates": self._dead_code(),
        }

    def _dead_code(self) -> List[Dict[str, str]]:
        """Symbols defined once and never mentioned anywhere else (best-effort)."""
        defined: Dict[str, str] = {}
        for rec in self.files:
            if rec.category not in ("Source",):
                continue
            for sym in rec.classes + rec.functions:
                if len(sym) < 5 or sym.startswith("_"):
                    continue
                defined.setdefault(sym, rec.rel_path)
        if not defined:
            return []
        counts: Counter = Counter()
        for rec in self.files:
            if rec.binary:
                continue
            text, is_bin = safe_read(os.path.join(self.root, rec.rel_path),
                                     min(self.max_bytes, 400000))
            if is_bin or not text:
                continue
            for sym in defined:
                if sym in text:
                    counts[sym] += text.count(sym)
        out = []
        for sym, path in sorted(defined.items()):
            if counts[sym] <= 1:
                out.append({"symbol": sym, "defined_in": path,
                            "note": "Defined once, no other textual reference found"})
        return out[:500]

    # ---------------------------------------------------------------- token math
    def token_analysis(self) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        prompts_by_file: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        for p in self.prompts:
            prompts_by_file[p["file"]].append(p)

        for a in sorted(self.agents, key=lambda x: (x["file"], x["name"])):
            params = self.params.get(a["file"], {})
            fprompts = prompts_by_file.get(a["file"], [])
            in_tokens = sum(p["estimated_tokens"] for p in fprompts)
            max_tokens = params.get("max_tokens", [])
            max_tok_val = max((int(v) for v in max_tokens), default=None)
            out_tokens = max_tok_val if max_tok_val else None
            model = a["models"][0] if a["models"] else None
            ctx = None
            if params.get("context_window"):
                ctx = int(params["context_window"][0])
            elif model:
                ctx = context_for(model)
            cost = NA
            if model and out_tokens is not None:
                pr = price_for(model)
                if pr:
                    c = (in_tokens / 1e6) * pr[0] + (out_tokens / 1e6) * pr[1]
                    cost = f"~${c:.6f} per request (list price estimate)"
            avg = (in_tokens + out_tokens) if (out_tokens is not None) else None
            rows.append({
                "agent": a["name"], "file": a["file"],
                "model": model or NA,
                "estimated_input_tokens": in_tokens if fprompts else 0,
                "estimated_input_tokens_note":
                    "Sum of static prompt text (~4 chars/token)" if fprompts
                    else "No static prompt text found in this file",
                "estimated_output_tokens": out_tokens if out_tokens is not None
                    else NA + " (no max_tokens configured)",
                "average_request_tokens": avg if avg is not None else NA,
                "max_tokens_configured": max_tok_val if max_tok_val is not None else NA,
                "temperature": params.get("temperature", [NA])[0],
                "top_p": params.get("top_p", [NA])[0],
                "context_window": ctx if ctx else NA,
                "estimated_cost": cost,
                "current_tokens_used": NA,
                "todays_tokens_used": NA,
                "monthly_tokens_used": NA,
                "remaining_tokens": NA,
                "percentage_used": NA,
            })
        return rows

    def request_analysis(self) -> Dict[str, Any]:
        agg: Dict[str, Set[str]] = defaultdict(set)
        for rel, params in sorted(self.params.items()):
            for key in ("requests_per_minute", "requests_per_day", "requests_per_month",
                        "tokens_per_minute", "concurrency", "timeout", "max_retries"):
                for v in params.get(key, []):
                    agg[key].add(f"{v} ({rel})")
        retry_files = sorted(self.workflows.get("Retry", []))
        backoff_files = sorted(self.workflows.get("Backoff", []))
        return {
            "requests_per_minute": sorted(agg.get("requests_per_minute", [])) or NA,
            "requests_per_day": sorted(agg.get("requests_per_day", [])) or NA,
            "requests_per_month": sorted(agg.get("requests_per_month", [])) or NA,
            "tokens_per_minute": sorted(agg.get("tokens_per_minute", [])) or NA,
            "max_concurrency": sorted(agg.get("concurrency", [])) or NA,
            "current_requests_used": NA,
            "remaining_requests": NA,
            "rate_limits_configured": sorted(
                set(list(agg.get("requests_per_minute", [])) +
                    list(agg.get("requests_per_day", [])) +
                    list(agg.get("tokens_per_minute", [])))) or NA,
            "retry_logic": retry_files or NA,
            "backoff_logic": backoff_files or NA,
            "timeout": sorted(agg.get("timeout", [])) or NA,
            "max_retries": sorted(agg.get("max_retries", [])) or NA,
            "runtime_usage_note":
                "Live request counters and remaining quota are runtime values held by the "
                "provider; they cannot be derived from source code.",
            "runtime_instrumentation_found":
                sorted({m["signal"] for m in self.runtime_metrics}) or
                ["No runtime usage instrumentation detected in source"],
        }

    # ---------------------------------------------------------------- stats
    def statistics(self) -> Dict[str, Any]:
        f = self.files
        lang_counter = Counter(r.language for r in f)
        cat_counter = Counter(r.category for r in f)
        purpose_counter = Counter(r.purpose for r in f)
        packages = {os.path.dirname(r.rel_path) for r in f if os.path.dirname(r.rel_path)}
        k8s = [r.rel_path for r in f if r.purpose == "Kubernetes"]
        docker = [r.rel_path for r in f
                  if r.language == "Dockerfile" or "docker-compose" in r.name.lower()]
        return {
            "total_directories": len(self.dirs),
            "total_files": len(f),
            "total_source_files": cat_counter.get("Source", 0),
            "total_configuration_files": cat_counter.get("Configuration", 0),
            "total_documentation_files": cat_counter.get("Documentation", 0),
            "total_test_files": cat_counter.get("Test", 0),
            "total_infrastructure_files": cat_counter.get("Infrastructure", 0),
            "total_ai_related_files": sum(1 for r in f if r.is_ai_related == "Yes"),
            "total_lines": sum(r.lines_total for r in f),
            "total_code_lines": sum(r.lines_code for r in f),
            "total_size_bytes": sum(r.size_bytes for r in f),
            "total_size_human": human_size(sum(r.size_bytes for r in f)),
            "total_classes": sum(len(r.classes) for r in f),
            "total_functions": sum(len(r.functions) for r in f),
            "total_apis": sum(1 for r in f if r.endpoints),
            "total_endpoints": sum(len(r.endpoints) for r in f),
            "total_modules": sum(1 for r in f if r.category == "Source"),
            "total_packages": len(packages),
            "total_tests": cat_counter.get("Test", 0),
            "total_docker_files": len(docker),
            "total_yaml_files": lang_counter.get("YAML", 0),
            "total_terraform_files": lang_counter.get("Terraform", 0) + lang_counter.get("HCL", 0),
            "total_kubernetes_files": len(k8s),
            "total_prompts": len(self.prompts),
            "total_models": len(self.models),
            "total_providers": len(self.providers),
            "total_agents": len(self.agents),
            "total_sdks": len(self.sdks),
            "total_tools": len(self.tools),
            "total_workflows": len(self.workflows),
            "total_env_vars": len(self.env_keys),
            "total_ai_env_vars": sum(1 for v in self.env_keys.values() if v["ai_related"]),
            "languages": dict(sorted(lang_counter.items(), key=lambda kv: (-kv[1], kv[0]))),
            "categories": dict(sorted(cat_counter.items(), key=lambda kv: (-kv[1], kv[0]))),
            "purposes": dict(sorted(purpose_counter.items(), key=lambda kv: (-kv[1], kv[0]))),
        }

    def directory_tree(self, max_entries: int = 4000) -> List[str]:
        lines = [os.path.basename(self.root) + "/"]
        tree: Dict[str, List[str]] = defaultdict(list)
        for r in self.files:
            tree[os.path.dirname(r.rel_path)].append(r.name)
        all_dirs = sorted(set(self.dirs) | {""})
        count = 0
        for d in all_dirs:
            depth = 0 if d == "" else d.count("/") + 1
            if d:
                lines.append("│   " * (depth - 1) + "├── " + os.path.basename(d) + "/")
            for fn in sorted(tree.get(d, [])):
                count += 1
                if count > max_entries:
                    lines.append("│   " * depth + "└── ... (truncated)")
                    return lines
                lines.append("│   " * depth + "├── " + fn)
        return lines

    # ---------------------------------------------------------------- assemble
    def build_report(self) -> Dict[str, Any]:
        graphs = self.build_graphs()
        stats = self.statistics()
        providers = []
        for name, d in sorted(self.providers.items()):
            providers.append({
                "provider": name, "sdk": d["sdk"],
                "endpoint": sorted(d["endpoints"]) or [NA],
                "api_version": sorted(d["api_versions"]) or [NA],
                "authentication_method": sorted(d["auth_methods"]) or [NA],
                "environment_variables": sorted(d["env_vars"]) or [NA],
                "files": sorted(d["files"]),
                "file_count": len(d["files"]),
            })
        models = []
        for name, d in sorted(self.models.items()):
            mv = re.search(r"[\d]{4}-[\d]{2}-[\d]{2}|v?\d+(?:\.\d+)*", name)
            models.append({
                "model": name, "version": mv.group(0) if mv else NA,
                "provider": d["provider"], "files": sorted(d["files"]),
                "reference_count": d["references"],
            })
        agents = []
        for a in sorted(self.agents, key=lambda x: (x["file"], x["name"])):
            agents.append({
                "name": a["name"], "type": a["type"], "file": a["file"],
                "purpose": a["purpose"],
                "providers": a["providers"] or [NA],
                "models": a["models"] or [NA],
                "sdks": a["sdks"] or [NA],
                "tools": a["tools"] or [],
                "workflows": a["workflows"] or [],
                "prompts": a["prompts"] or [],
                "env_vars": a["env_vars"] or [],
                "ai_call_sites": a["ai_calls"],
            })
        report: Dict[str, Any] = {
            "meta": {
                "tool": "AI Repository Auditor",
                "version": VERSION,
                "analysis_mode": "Static analysis only (no repository code executed, "
                                 "no files modified)",
                "token_estimation_method": "Character heuristic (~4 characters per token)",
                "unavailable_marker": NA,
            },
            "repository": {
                "repository_name": os.path.basename(self.root) or self.root,
                "repository_root": self.root,
                "scan_date": self.scan_started.strftime("%Y-%m-%d %H:%M:%S"),
                "scan_duration_seconds": round(
                    (_dt.datetime.now() - self.scan_started).total_seconds(), 2),
                "total_directories": stats["total_directories"],
                "total_files": stats["total_files"],
                "total_source_files": stats["total_source_files"],
                "total_configuration_files": stats["total_configuration_files"],
                "total_documentation_files": stats["total_documentation_files"],
                "total_test_files": stats["total_test_files"],
                "total_infrastructure_files": stats["total_infrastructure_files"],
                "total_ai_related_files": stats["total_ai_related_files"],
                "total_size": stats["total_size_human"],
                "total_lines": stats["total_lines"],
            },
            "statistics": stats,
            "directory_tree": self.directory_tree(),
            "file_inventory": [self._file_row(r) for r in self.files],
            "ai_agents": {"total_ai_agents": len(agents), "agents": agents},
            "ai_providers": providers,
            "ai_models": models,
            "prompts": sorted(self.prompts, key=lambda p: (p["file"], p["name"])),
            "api_keys": [
                {"variable": k, "provider": v["provider"],
                 "used_in": sorted(v["files"]),
                 "loaded_from": sorted(v["loaded_from"]) or [NA],
                 "ai_related": "Yes" if v["ai_related"] else "No"}
                for k, v in sorted(self.env_keys.items())],
            "token_usage": {
                "per_agent": self.token_analysis(),
                "runtime_usage": {
                    "current_tokens_used": NA, "todays_tokens_used": NA,
                    "monthly_tokens_used": NA, "remaining_tokens": NA,
                    "percentage_used": NA,
                    "note": "Runtime token counters live in the provider dashboard / "
                            "telemetry backend and are not present in source code.",
                    "instrumentation_found":
                        sorted({m["signal"] for m in self.runtime_metrics}) or
                        ["No token-usage instrumentation detected in source"],
                    "instrumentation_files":
                        sorted({m["file"] for m in self.runtime_metrics}) or [],
                },
            },
            "request_usage": self.request_analysis(),
            "ai_sdks": [{"sdk": k, "files": sorted(set(v)), "file_count": len(set(v))}
                        for k, v in sorted(self.sdks.items())],
            "tools": [{"tool": k, "files": sorted(set(v)), "file_count": len(set(v))}
                      for k, v in sorted(self.tools.items())],
            "workflows": [{"pattern": k, "files": sorted(set(v)), "file_count": len(set(v))}
                          for k, v in sorted(self.workflows.items())],
            "model_parameters": {k: v for k, v in sorted(self.params.items())},
            "dependencies": graphs,
            "scan_errors": self.errors,
        }
        return report

    @staticmethod
    def _file_row(r: FileRecord) -> Dict[str, Any]:
        d = asdict(r)
        d.pop("sha1", None)
        return d


# =====================================================================================
# SECTION 8 -- REPORT WRITERS
# =====================================================================================


def md_table(headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> str:
    def cell(v: Any) -> str:
        s = ", ".join(str(x) for x in v) if isinstance(v, (list, tuple, set)) else str(v)
        s = s.replace("|", "\\|").replace("\n", " ").strip()
        return s if s else "-"
    out = ["| " + " | ".join(headers) + " |",
           "|" + "|".join(["---"] * len(headers)) + "|"]
    for row in rows:
        out.append("| " + " | ".join(cell(c) for c in row) + " |")
    if not rows:
        out.append("| " + " | ".join(["-"] * len(headers)) + " |")
    return "\n".join(out)


def write_markdown(report: Dict[str, Any], path: str) -> None:
    R, S = report["repository"], report["statistics"]
    L: List[str] = []
    A = L.append

    A(f"# AI Repository Audit — {R['repository_name']}\n")
    A(f"> Generated by **AI Repository Auditor v{VERSION}** · "
      f"{report['meta']['analysis_mode']}\n")
    A("## 1. Repository Summary\n")
    A(md_table(["Metric", "Value"], [
        ["Repository Name", R["repository_name"]],
        ["Repository Root", R["repository_root"]],
        ["Scan Date", R["scan_date"]],
        ["Scan Duration", f"{R['scan_duration_seconds']} s"],
        ["Total Directories", R["total_directories"]],
        ["Total Files", R["total_files"]],
        ["Total Source Files", R["total_source_files"]],
        ["Total Configuration Files", R["total_configuration_files"]],
        ["Total Documentation Files", R["total_documentation_files"]],
        ["Total Test Files", R["total_test_files"]],
        ["Total Infrastructure Files", R["total_infrastructure_files"]],
        ["Total AI Related Files", R["total_ai_related_files"]],
        ["Total Size", R["total_size"]],
        ["Total Lines", R["total_lines"]],
    ]))

    A("\n## 2. Repository Statistics\n")
    A(md_table(["Statistic", "Count"], [
        ["Total Classes", S["total_classes"]], ["Total Functions", S["total_functions"]],
        ["Total APIs (files exposing endpoints)", S["total_apis"]],
        ["Total Endpoints", S["total_endpoints"]],
        ["Total Modules", S["total_modules"]], ["Total Packages", S["total_packages"]],
        ["Total Tests", S["total_tests"]], ["Total Docker Files", S["total_docker_files"]],
        ["Total YAML Files", S["total_yaml_files"]],
        ["Total Terraform Files", S["total_terraform_files"]],
        ["Total Kubernetes Files", S["total_kubernetes_files"]],
        ["Total Prompts", S["total_prompts"]], ["Total Models", S["total_models"]],
        ["Total Providers", S["total_providers"]], ["Total Agents", S["total_agents"]],
        ["Total SDKs", S["total_sdks"]], ["Total Tools", S["total_tools"]],
        ["Total Workflow Patterns", S["total_workflows"]],
        ["Total Environment Variables", S["total_env_vars"]],
        ["AI Environment Variables", S["total_ai_env_vars"]],
    ]))

    A("\n### Languages\n")
    A(md_table(["Language", "Files"], sorted(S["languages"].items(),
                                             key=lambda kv: (-kv[1], kv[0]))))
    A("\n### File Purposes\n")
    A(md_table(["Purpose", "Files"], sorted(S["purposes"].items(),
                                            key=lambda kv: (-kv[1], kv[0]))))

    A("\n## 3. Directory Tree\n")
    A("```\n" + "\n".join(report["directory_tree"][:1500]) + "\n```")

    A("\n## 4. File Inventory\n")
    rows = [[r["serial"], r["rel_path"], r["name"], r["ext"], r["language"], r["category"],
             r["purpose"], r["description"][:220], r["size_human"], r["lines_code"],
             r["is_used"], r["is_referenced"], r["is_ai_related"],
             ", ".join(r["dependencies"][:6]) or "-"] for r in report["file_inventory"]]
    A(md_table(["#", "Relative Path", "File Name", "Ext", "Language", "Category", "Purpose",
                "Description", "Size", "LOC", "Used", "Referenced", "AI Related",
                "Dependencies"], rows))

    A("\n## 5. AI Agents\n")
    A(f"**Total AI Agents: {report['ai_agents']['total_ai_agents']}**\n")
    A(md_table(["Agent Name", "Type", "File", "Purpose", "Provider", "Model", "SDK",
                "Tools", "Workflows", "Prompts"],
               [[a["name"], a["type"], a["file"], a["purpose"], a["providers"],
                 a["models"], a["sdks"], a["tools"], a["workflows"], a["prompts"]]
                for a in report["ai_agents"]["agents"]]))

    A("\n## 6. AI Providers\n")
    A(md_table(["Provider", "SDK", "Endpoint", "API Version", "Authentication",
                "Environment Variables", "Files"],
               [[p["provider"], p["sdk"], p["endpoint"][:3], p["api_version"][:3],
                 p["authentication_method"], p["environment_variables"][:6],
                 p["files"][:6]] for p in report["ai_providers"]]))

    A("\n## 7. AI Models\n")
    A(md_table(["Model", "Version", "Provider", "Files Using It", "References"],
               [[m["model"], m["version"], m["provider"], m["files"][:6],
                 m["reference_count"]] for m in report["ai_models"]]))

    A("\n## 8. Prompt Inventory\n")
    A(md_table(["Prompt Name", "Type", "Location", "Purpose", "Agent Using It",
                "Est. Tokens", "Preview"],
               [[p["name"], p["type"], f"{p['file']}:{p['line']}", p["purpose"],
                 p["agent"], p["estimated_tokens"], p["preview"][:110]]
                for p in report["prompts"]]))

    A("\n## 9. API Keys / Environment Variables\n")
    A(md_table(["Variable Name", "Provider", "AI Related", "Used In", "Loaded From"],
               [[k["variable"], k["provider"], k["ai_related"], k["used_in"][:5],
                 k["loaded_from"]] for k in report["api_keys"]]))

    A("\n## 10. Token Usage Analysis (static estimates)\n")
    A(md_table(["Agent", "File", "Model", "Est. Input Tokens", "Est. Output Tokens",
                "Avg Request Tokens", "Max Tokens", "Temperature", "Top P",
                "Context Window", "Estimated Cost"],
               [[t["agent"], t["file"], t["model"], t["estimated_input_tokens"],
                 t["estimated_output_tokens"], t["average_request_tokens"],
                 t["max_tokens_configured"], t["temperature"], t["top_p"],
                 t["context_window"], t["estimated_cost"]]
                for t in report["token_usage"]["per_agent"]]))
    ru = report["token_usage"]["runtime_usage"]
    A("\n**Runtime token usage**\n")
    A(md_table(["Metric", "Value"], [
        ["Current Tokens Used", ru["current_tokens_used"]],
        ["Today's Tokens Used", ru["todays_tokens_used"]],
        ["Monthly Tokens Used", ru["monthly_tokens_used"]],
        ["Remaining Tokens", ru["remaining_tokens"]],
        ["Percentage Used", ru["percentage_used"]],
        ["Instrumentation Found", ", ".join(ru["instrumentation_found"])],
    ]))
    A(f"\n_{ru['note']}_\n")

    A("\n## 11. Request Usage Analysis\n")
    q = report["request_usage"]
    A(md_table(["Metric", "Value"], [
        ["Requests Per Minute", q["requests_per_minute"]],
        ["Requests Per Day", q["requests_per_day"]],
        ["Requests Per Month", q["requests_per_month"]],
        ["Tokens Per Minute", q["tokens_per_minute"]],
        ["Max Concurrency", q["max_concurrency"]],
        ["Current Requests Used", q["current_requests_used"]],
        ["Remaining Requests", q["remaining_requests"]],
        ["Rate Limits", q["rate_limits_configured"]],
        ["Retry Logic", q["retry_logic"]],
        ["Backoff Logic", q["backoff_logic"]],
        ["Timeout", q["timeout"]],
        ["Max Retries", q["max_retries"]],
    ]))
    A(f"\n_{q['runtime_usage_note']}_\n")

    A("\n## 12. AI SDKs / Frameworks\n")
    A(md_table(["SDK / Framework", "Files", "File Count"],
               [[s["sdk"], s["files"][:8], s["file_count"]] for s in report["ai_sdks"]]))

    A("\n## 13. Tools Used by Agents\n")
    A(md_table(["Tool", "Files", "File Count"],
               [[t["tool"], t["files"][:8], t["file_count"]] for t in report["tools"]]))

    A("\n## 14. Agentic Workflow Patterns\n")
    A(md_table(["Pattern", "Files", "File Count"],
               [[w["pattern"], w["files"][:8], w["file_count"]] for w in report["workflows"]]))

    A("\n## 15. Dependencies\n")
    dep = report["dependencies"]
    A("### File Dependency Graph (internal imports)\n")
    A(md_table(["File", "Depends On"],
               [[k, v[:10]] for k, v in sorted(dep["file_dependency_graph"].items())]))
    A("\n### Agent Dependency Graph\n")
    A(md_table(["Agent", "File", "Providers", "Models", "SDKs", "Tools"],
               [[a["agent"], a["file"], a["providers"], a["models"], a["sdks"], a["tools"]]
                for a in dep["agent_dependency_graph"]]))
    A("\n### Model Dependency Graph\n")
    A(md_table(["Model", "Provider", "Files"],
               [[m["model"], m["provider"], m["files"][:8]]
                for m in dep["model_dependency_graph"]]))
    A("\n### Unused / Orphan Files\n")
    A(md_table(["File"], [[f] for f in dep["orphan_files"]]))
    A("\n### Dead Code Candidates\n")
    A(md_table(["Symbol", "Defined In", "Note"],
               [[d["symbol"], d["defined_in"], d["note"]]
                for d in dep["dead_code_candidates"][:200]]))

    A("\n---\n")
    A(f"_Report generated {R['scan_date']} · static analysis only · "
      f"values that cannot be derived from source are reported as "
      f"\"{NA}\"._\n")

    with open(path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(L))


def write_json(report: Dict[str, Any], path: str) -> None:
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(report, fh, indent=2, sort_keys=False, default=str)


CSV_COLUMNS = ["serial", "rel_path", "name", "ext", "language", "category", "purpose",
               "description", "size_bytes", "size_human", "lines_total", "lines_code",
               "is_used", "is_referenced", "is_ai_related", "dependencies",
               "internal_dependencies", "classes", "functions", "endpoints", "providers",
               "models", "sdks", "tools", "workflows", "env_vars", "agents",
               "prompt_count", "ai_calls"]


def write_csv(report: Dict[str, Any], path: str) -> None:
    with open(path, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow([c.replace("_", " ").title() for c in CSV_COLUMNS])
        for r in report["file_inventory"]:
            row = []
            for c in CSV_COLUMNS:
                v = r.get(c, "")
                row.append("; ".join(str(x) for x in v) if isinstance(v, list) else v)
            w.writerow(row)


def write_xlsx(report: Dict[str, Any], path: str) -> Optional[str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return None

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(color="FFFFFF", bold=True)

    def sheet(title: str, headers: Sequence[str], rows: Sequence[Sequence[Any]],
              first: bool = False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title[:31]
        ws.append(list(headers))
        for c in ws[1]:
            c.fill, c.font = head_fill, head_font
            c.alignment = Alignment(vertical="center")
        for row in rows:
            ws.append(["; ".join(str(x) for x in v) if isinstance(v, (list, tuple, set))
                       else v for v in row])
        for i, h in enumerate(headers, 1):
            width = max(len(str(h)) + 2,
                        *(len(str(r[i - 1])[:60]) + 2 for r in rows[:300])) if rows \
                    else len(str(h)) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(max(width, 10), 60)
        ws.freeze_panes = "A2"
        if rows:
            ws.auto_filter.ref = ws.dimensions
        return ws

    R, S = report["repository"], report["statistics"]
    sheet("Summary", ["Metric", "Value"],
          [[k.replace("_", " ").title(), v] for k, v in R.items()] +
          [["", ""]] +
          [[k.replace("_", " ").title(), v] for k, v in S.items()
           if not isinstance(v, dict)], first=True)
    sheet("File Inventory", [c.replace("_", " ").title() for c in CSV_COLUMNS],
          [[r.get(c, "") for c in CSV_COLUMNS] for r in report["file_inventory"]])
    sheet("AI Agents", ["Name", "Type", "File", "Purpose", "Providers", "Models", "SDKs",
                        "Tools", "Workflows", "Prompts", "AI Call Sites"],
          [[a["name"], a["type"], a["file"], a["purpose"], a["providers"], a["models"],
            a["sdks"], a["tools"], a["workflows"], a["prompts"], a["ai_call_sites"]]
           for a in report["ai_agents"]["agents"]])
    sheet("Providers", ["Provider", "SDK", "Endpoint", "API Version", "Authentication",
                        "Env Vars", "Files"],
          [[p["provider"], p["sdk"], p["endpoint"], p["api_version"],
            p["authentication_method"], p["environment_variables"], p["files"]]
           for p in report["ai_providers"]])
    sheet("Models", ["Model", "Version", "Provider", "Files", "References"],
          [[m["model"], m["version"], m["provider"], m["files"], m["reference_count"]]
           for m in report["ai_models"]])
    sheet("Prompts", ["Name", "Type", "File", "Line", "Purpose", "Agent", "Chars",
                      "Est Tokens", "Preview"],
          [[p["name"], p["type"], p["file"], p["line"], p["purpose"], p["agent"],
            p["chars"], p["estimated_tokens"], p["preview"]] for p in report["prompts"]])
    sheet("API Keys", ["Variable", "Provider", "AI Related", "Used In", "Loaded From"],
          [[k["variable"], k["provider"], k["ai_related"], k["used_in"], k["loaded_from"]]
           for k in report["api_keys"]])
    sheet("Token Usage", ["Agent", "File", "Model", "Est Input Tokens", "Est Output Tokens",
                          "Avg Request Tokens", "Max Tokens", "Temperature", "Top P",
                          "Context Window", "Estimated Cost", "Current Tokens Used",
                          "Today Tokens Used", "Monthly Tokens Used", "Remaining Tokens",
                          "Percentage Used"],
          [[t["agent"], t["file"], t["model"], t["estimated_input_tokens"],
            t["estimated_output_tokens"], t["average_request_tokens"],
            t["max_tokens_configured"], t["temperature"], t["top_p"], t["context_window"],
            t["estimated_cost"], t["current_tokens_used"], t["todays_tokens_used"],
            t["monthly_tokens_used"], t["remaining_tokens"], t["percentage_used"]]
           for t in report["token_usage"]["per_agent"]])
    q = report["request_usage"]
    sheet("Request Usage", ["Metric", "Value"],
          [[k.replace("_", " ").title(), v] for k, v in q.items()])
    sheet("SDKs", ["SDK", "File Count", "Files"],
          [[s["sdk"], s["file_count"], s["files"]] for s in report["ai_sdks"]])
    sheet("Tools", ["Tool", "File Count", "Files"],
          [[t["tool"], t["file_count"], t["files"]] for t in report["tools"]])
    sheet("Workflows", ["Pattern", "File Count", "Files"],
          [[w["pattern"], w["file_count"], w["files"]] for w in report["workflows"]])
    dep = report["dependencies"]
    sheet("Dependencies", ["File", "Depends On"],
          [[k, v] for k, v in sorted(dep["file_dependency_graph"].items())])
    sheet("Unused Files", ["File"], [[f] for f in dep["orphan_files"]])
    sheet("Dead Code", ["Symbol", "Defined In", "Note"],
          [[d["symbol"], d["defined_in"], d["note"]] for d in dep["dead_code_candidates"]])
    wb.save(path)
    return path


# =====================================================================================
# SECTION 9 -- HTML DASHBOARD
# =====================================================================================

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>AI Repository Audit — __REPO__</title>
<style>
:root{--bg:#0e1117;--panel:#161b22;--panel2:#1c2230;--txt:#e6edf3;--muted:#8b949e;
--acc:#4f9cf9;--acc2:#7ee787;--warn:#f0883e;--bad:#f85149;--brd:#30363d;}
html[data-theme="light"]{--bg:#f5f7fa;--panel:#ffffff;--panel2:#eef2f7;--txt:#1b2430;
--muted:#5b6673;--acc:#1a63d8;--acc2:#1a7f37;--warn:#b45309;--bad:#b91c1c;--brd:#d5dbe3;}
*{box-sizing:border-box}
body{margin:0;background:var(--bg);color:var(--txt);
font:14px/1.5 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif}
header{position:sticky;top:0;z-index:50;background:var(--panel);border-bottom:1px solid var(--brd);
padding:14px 20px;display:flex;flex-wrap:wrap;gap:12px;align-items:center;justify-content:space-between}
h1{font-size:18px;margin:0}
h2{font-size:17px;margin:0 0 12px;padding-bottom:8px;border-bottom:1px solid var(--brd)}
h3{font-size:14px;margin:18px 0 8px;color:var(--muted);text-transform:uppercase;letter-spacing:.06em}
.sub{color:var(--muted);font-size:12px}
.btn{background:var(--panel2);color:var(--txt);border:1px solid var(--brd);border-radius:8px;
padding:7px 12px;cursor:pointer;font-size:13px}
.btn:hover{border-color:var(--acc);color:var(--acc)}
.wrap{max-width:1680px;margin:0 auto;padding:20px}
nav{display:flex;flex-wrap:wrap;gap:6px;margin-bottom:18px}
nav a{padding:6px 11px;border-radius:20px;background:var(--panel);border:1px solid var(--brd);
color:var(--muted);text-decoration:none;font-size:12.5px}
nav a:hover{color:var(--acc);border-color:var(--acc)}
section{background:var(--panel);border:1px solid var(--brd);border-radius:12px;
padding:18px;margin-bottom:18px}
.cards{display:grid;grid-template-columns:repeat(auto-fill,minmax(165px,1fr));gap:12px}
.card{background:var(--panel2);border:1px solid var(--brd);border-radius:10px;padding:14px}
.card .v{font-size:24px;font-weight:700;color:var(--acc)}
.card .l{font-size:11.5px;color:var(--muted);text-transform:uppercase;letter-spacing:.05em}
.tblwrap{overflow:auto;max-height:640px;border:1px solid var(--brd);border-radius:10px}
table{border-collapse:collapse;width:100%;font-size:12.5px}
th,td{padding:7px 9px;border-bottom:1px solid var(--brd);text-align:left;vertical-align:top}
th{background:var(--panel2);position:sticky;top:0;cursor:pointer;white-space:nowrap;z-index:2}
th:hover{color:var(--acc)}
tbody tr:hover{background:rgba(79,156,249,.08)}
td.desc{max-width:420px}
input,select{background:var(--panel2);color:var(--txt);border:1px solid var(--brd);
border-radius:8px;padding:7px 10px;font-size:13px}
.controls{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:10px;align-items:center}
.tag{display:inline-block;background:var(--panel2);border:1px solid var(--brd);border-radius:6px;
padding:1px 7px;margin:1px 3px 1px 0;font-size:11.5px}
.yes{color:var(--acc2);font-weight:600}.no{color:var(--muted)}.na{color:var(--warn)}
pre{background:var(--panel2);border:1px solid var(--brd);border-radius:10px;padding:14px;
overflow:auto;max-height:520px;font-size:12px;margin:0}
.charts{display:grid;grid-template-columns:repeat(auto-fit,minmax(330px,1fr));gap:16px}
.chart{background:var(--panel2);border:1px solid var(--brd);border-radius:10px;padding:14px}
.chart h4{margin:0 0 10px;font-size:13px;color:var(--muted)}
.bar{display:flex;align-items:center;gap:8px;margin:5px 0}
.bar .nm{width:38%;font-size:11.5px;color:var(--muted);overflow:hidden;text-overflow:ellipsis;
white-space:nowrap}
.bar .tr{flex:1;background:var(--bg);border-radius:5px;height:15px;overflow:hidden}
.bar .fl{height:100%;background:linear-gradient(90deg,var(--acc),var(--acc2));border-radius:5px}
.bar .ct{width:52px;text-align:right;font-size:11.5px;color:var(--txt)}
.note{background:rgba(240,136,62,.1);border-left:3px solid var(--warn);padding:10px 12px;
border-radius:0 8px 8px 0;font-size:12.5px;margin:10px 0}
footer{color:var(--muted);font-size:12px;text-align:center;padding:22px}
@media(max-width:720px){.wrap{padding:12px}header{padding:10px 12px}td.desc{max-width:200px}}
</style>
</head>
<body>
<header>
  <div><h1>AI Repository Audit — __REPO__</h1>
  <div class="sub">__ROOT__ · scanned __DATE__ · static analysis only</div></div>
  <div style="display:flex;gap:8px;flex-wrap:wrap">
    <button class="btn" onclick="toggleTheme()">🌓 Theme</button>
    <button class="btn" onclick="exportJSON()">⬇ JSON</button>
    <button class="btn" onclick="exportCSV()">⬇ CSV (inventory)</button>
    <button class="btn" onclick="window.print()">🖨 Print / PDF</button>
  </div>
</header>
<div class="wrap">
<nav>
  <a href="#overview">Overview</a><a href="#stats">File Statistics</a>
  <a href="#charts">Charts</a><a href="#tree">Directory Tree</a>
  <a href="#files">File Purpose Table</a><a href="#agents">AI Agents</a>
  <a href="#providers">Providers</a><a href="#models">Models</a>
  <a href="#prompts">Prompts</a><a href="#tokens">Token Usage</a>
  <a href="#requests">Request Usage</a><a href="#sdks">SDKs</a>
  <a href="#tools">Tools &amp; Workflows</a><a href="#deps">Dependencies</a>
  <a href="#unused">Unused Files</a>
</nav>

<section id="overview"><h2>Repository Overview</h2><div class="cards" id="ovCards"></div></section>
<section id="stats"><h2>File Statistics</h2><div class="cards" id="stCards"></div></section>
<section id="charts"><h2>Charts</h2><div class="charts" id="chartArea"></div></section>
<section id="tree"><h2>Directory Tree</h2><pre id="tree_pre"></pre></section>

<section id="files"><h2>File Purpose Table</h2>
  <div class="controls">
    <input id="q" placeholder="Search path, purpose, description…" style="min-width:280px;flex:1"/>
    <select id="fCat"></select><select id="fLang"></select>
    <select id="fPurpose"></select><select id="fAI"></select>
    <button class="btn" onclick="resetFilters()">Reset</button>
    <span class="sub" id="cnt"></span>
  </div>
  <div class="tblwrap"><table id="tFiles"><thead></thead><tbody></tbody></table></div>
</section>

<section id="agents"><h2>AI Agents <span class="sub">(total: __AGENTS__)</span></h2>
  <div class="tblwrap"><table id="tAgents"><thead></thead><tbody></tbody></table></div></section>
<section id="providers"><h2>AI Providers</h2>
  <div class="tblwrap"><table id="tProv"><thead></thead><tbody></tbody></table></div></section>
<section id="models"><h2>AI Models</h2>
  <div class="tblwrap"><table id="tModels"><thead></thead><tbody></tbody></table></div></section>
<section id="prompts"><h2>Prompt Inventory</h2>
  <div class="controls"><input id="qp" placeholder="Search prompts…" style="min-width:280px;flex:1"/></div>
  <div class="tblwrap"><table id="tPrompts"><thead></thead><tbody></tbody></table></div></section>
<section id="keys"><h2>API Keys / Environment Variables</h2>
  <div class="tblwrap"><table id="tKeys"><thead></thead><tbody></tbody></table></div></section>

<section id="tokens"><h2>Token Usage</h2>
  <div class="note">Token figures below are <b>static estimates</b> derived from prompt text
  (~4 characters per token) and configured <code>max_tokens</code>. Live consumption is a
  runtime value and is reported as “__NA__”.</div>
  <div class="tblwrap"><table id="tTokens"><thead></thead><tbody></tbody></table></div>
  <h3>Runtime Token Usage</h3>
  <div class="tblwrap"><table id="tRuntime"><thead></thead><tbody></tbody></table></div>
</section>

<section id="requests"><h2>Request Usage</h2>
  <div class="tblwrap"><table id="tReq"><thead></thead><tbody></tbody></table></div></section>
<section id="sdks"><h2>AI SDKs / Frameworks</h2>
  <div class="tblwrap"><table id="tSdks"><thead></thead><tbody></tbody></table></div></section>
<section id="tools"><h2>Tools &amp; Workflow Patterns</h2>
  <h3>Tools</h3><div class="tblwrap"><table id="tTools"><thead></thead><tbody></tbody></table></div>
  <h3>Workflows</h3><div class="tblwrap"><table id="tWf"><thead></thead><tbody></tbody></table></div>
</section>
<section id="deps"><h2>Dependencies</h2>
  <h3>File Dependency Graph</h3>
  <div class="tblwrap"><table id="tDeps"><thead></thead><tbody></tbody></table></div>
  <h3>Agent Dependency Graph</h3>
  <div class="tblwrap"><table id="tADeps"><thead></thead><tbody></tbody></table></div>
</section>
<section id="unused"><h2>Unused / Orphan Files &amp; Dead Code</h2>
  <h3>Unused / Orphan Files</h3>
  <div class="tblwrap"><table id="tOrph"><thead></thead><tbody></tbody></table></div>
  <h3>Dead Code Candidates</h3>
  <div class="tblwrap"><table id="tDead"><thead></thead><tbody></tbody></table></div>
</section>

<footer>AI Repository Auditor v__VERSION__ · deterministic static analysis ·
no repository code was executed and no files were modified.</footer>
</div>

<script id="auditdata" type="application/json">__DATA__</script>
<script>
const DATA = JSON.parse(document.getElementById('auditdata').textContent);
const NA = "__NA__";
const esc = s => String(s==null?'':s).replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
const tags = a => (Array.isArray(a)?a:[a]).filter(x=>x!==''&&x!=null)
    .map(x=>`<span class="tag">${esc(x)}</span>`).join('') || '<span class="no">—</span>';
const flag = v => v==='Yes' ? '<span class="yes">Yes</span>'
    : (String(v).startsWith('Yes')||String(v).startsWith('Likely')) ? `<span class="yes">${esc(v)}</span>`
    : String(v)===NA ? `<span class="na">${esc(v)}</span>` : `<span class="no">${esc(v)}</span>`;

function card(l,v){return `<div class="card"><div class="v">${esc(v)}</div><div class="l">${esc(l)}</div></div>`;}
function fillCards(id, obj){document.getElementById(id).innerHTML =
  Object.entries(obj).map(([k,v])=>card(k,v)).join('');}

function renderTable(id, headers, rows, opts){
  opts = opts||{};
  const t = document.getElementById(id);
  t.querySelector('thead').innerHTML = '<tr>'+headers.map((h,i)=>
     `<th onclick="sortTable('${id}',${i})">${esc(h)} ⇅</th>`).join('')+'</tr>';
  t.querySelector('tbody').innerHTML = rows.length ? rows.map(r=>'<tr>'+r.map((c,i)=>
     `<td class="${opts.descCols&&opts.descCols.includes(i)?'desc':''}">${c}</td>`).join('')+'</tr>').join('')
     : `<tr><td colspan="${headers.length}" class="no">No entries detected.</td></tr>`;
}
function sortTable(id, col){
  const tb = document.querySelector('#'+id+' tbody');
  const rows = Array.from(tb.rows);
  const dir = tb.dataset.sc==String(col) && tb.dataset.sd=='asc' ? 'desc':'asc';
  tb.dataset.sc = col; tb.dataset.sd = dir;
  rows.sort((a,b)=>{
    const x=a.cells[col]?.innerText.trim()||'', y=b.cells[col]?.innerText.trim()||'';
    const nx=parseFloat(x.replace(/[^0-9.\-]/g,'')), ny=parseFloat(y.replace(/[^0-9.\-]/g,''));
    const both = !isNaN(nx)&&!isNaN(ny)&&x.match(/\d/)&&y.match(/\d/);
    const c = both ? nx-ny : x.localeCompare(y);
    return dir=='asc'?c:-c;});
  rows.forEach(r=>tb.appendChild(r));
}

/* ---------- overview + stats ---------- */
const R=DATA.repository, S=DATA.statistics;
fillCards('ovCards',{ 'Directories':R.total_directories,'Files':R.total_files,
 'Source Files':R.total_source_files,'Config Files':R.total_configuration_files,
 'Docs':R.total_documentation_files,'Tests':R.total_test_files,
 'Infra Files':R.total_infrastructure_files,'AI Files':R.total_ai_related_files,
 'Total Size':R.total_size,'Total Lines':R.total_lines});
fillCards('stCards',{'Classes':S.total_classes,'Functions':S.total_functions,
 'APIs':S.total_apis,'Endpoints':S.total_endpoints,'Modules':S.total_modules,
 'Packages':S.total_packages,'Docker':S.total_docker_files,'YAML':S.total_yaml_files,
 'Terraform':S.total_terraform_files,'Kubernetes':S.total_kubernetes_files,
 'Prompts':S.total_prompts,'Models':S.total_models,'Providers':S.total_providers,
 'Agents':S.total_agents,'SDKs':S.total_sdks,'Env Vars':S.total_env_vars});

/* ---------- charts ---------- */
function barChart(title, entries){
  entries = entries.slice(0,12);
  const max = Math.max(1,...entries.map(e=>e[1]));
  return `<div class="chart"><h4>${esc(title)}</h4>`+entries.map(([k,v])=>
    `<div class="bar"><div class="nm" title="${esc(k)}">${esc(k)}</div>
     <div class="tr"><div class="fl" style="width:${(v/max*100).toFixed(1)}%"></div></div>
     <div class="ct">${v}</div></div>`).join('')+'</div>';
}
const providerCounts = DATA.ai_providers.map(p=>[p.provider,p.file_count]).sort((a,b)=>b[1]-a[1]);
const modelCounts = DATA.ai_models.map(m=>[m.model,m.reference_count]).sort((a,b)=>b[1]-a[1]);
const sdkCounts = DATA.ai_sdks.map(s=>[s.sdk,s.file_count]).sort((a,b)=>b[1]-a[1]);
const locByLang = {};
DATA.file_inventory.forEach(f=>{locByLang[f.language]=(locByLang[f.language]||0)+f.lines_code;});
document.getElementById('chartArea').innerHTML =
  barChart('Files by Language', Object.entries(S.languages))
+ barChart('Files by Purpose', Object.entries(S.purposes))
+ barChart('Files by Category', Object.entries(S.categories))
+ barChart('Lines of Code by Language', Object.entries(locByLang).sort((a,b)=>b[1]-a[1]))
+ barChart('AI Providers (files)', providerCounts)
+ barChart('AI Models (references)', modelCounts)
+ barChart('AI SDKs (files)', sdkCounts)
+ barChart('Agents per File', Object.entries(DATA.ai_agents.agents.reduce((a,x)=>
    (a[x.file]=(a[x.file]||0)+1,a),{})).sort((a,b)=>b[1]-a[1]));

document.getElementById('tree_pre').textContent = DATA.directory_tree.join('\n');

/* ---------- file table with search / filter ---------- */
const FH = ['#','Path','Name','Ext','Language','Category','Purpose','Description','Size',
            'LOC','Used','Referenced','AI','Dependencies'];
function fileRow(f){return [f.serial, esc(f.rel_path), esc(f.name), esc(f.ext),
  esc(f.language), esc(f.category), `<b>${esc(f.purpose)}</b>`, esc(f.description),
  esc(f.size_human), f.lines_code, flag(f.is_used), flag(f.is_referenced),
  flag(f.is_ai_related), tags((f.dependencies||[]).slice(0,6))];}
function opts(id,label,vals){const s=document.getElementById(id);
  s.innerHTML = `<option value="">${label}: All</option>`+
  [...new Set(vals)].sort().map(v=>`<option>${esc(v)}</option>`).join('');
  s.onchange = applyFilters;}
opts('fCat','Category',DATA.file_inventory.map(f=>f.category));
opts('fLang','Language',DATA.file_inventory.map(f=>f.language));
opts('fPurpose','Purpose',DATA.file_inventory.map(f=>f.purpose));
opts('fAI','AI',DATA.file_inventory.map(f=>f.is_ai_related));
function applyFilters(){
  const q=document.getElementById('q').value.toLowerCase();
  const c=document.getElementById('fCat').value, l=document.getElementById('fLang').value;
  const p=document.getElementById('fPurpose').value, a=document.getElementById('fAI').value;
  const rows=DATA.file_inventory.filter(f=>
    (!c||f.category===c)&&(!l||f.language===l)&&(!p||f.purpose===p)&&(!a||f.is_ai_related===a)&&
    (!q||(f.rel_path+' '+f.purpose+' '+f.description+' '+f.language+' '+
          (f.agents||[]).join(' ')+' '+(f.models||[]).join(' ')).toLowerCase().includes(q)));
  renderTable('tFiles',FH,rows.map(fileRow),{descCols:[7]});
  document.getElementById('cnt').textContent = rows.length+' / '+DATA.file_inventory.length+' files';
}
document.getElementById('q').oninput = applyFilters;
function resetFilters(){['q','fCat','fLang','fPurpose','fAI'].forEach(i=>document.getElementById(i).value='');applyFilters();}
applyFilters();

/* ---------- AI sections ---------- */
renderTable('tAgents',['Agent','Type','File','Purpose','Providers','Models','SDKs','Tools','Workflows','Prompts','AI Calls'],
  DATA.ai_agents.agents.map(a=>[esc(a.name),esc(a.type),esc(a.file),esc(a.purpose),
    tags(a.providers),tags(a.models),tags(a.sdks),tags(a.tools),tags(a.workflows),
    tags(a.prompts),a.ai_call_sites]),{descCols:[3]});
renderTable('tProv',['Provider','SDK','Endpoint','API Version','Authentication','Env Vars','Files'],
  DATA.ai_providers.map(p=>[esc(p.provider),esc(p.sdk),tags(p.endpoint),tags(p.api_version),
    tags(p.authentication_method),tags(p.environment_variables),tags(p.files)]));
renderTable('tModels',['Model','Version','Provider','Files Using It','References'],
  DATA.ai_models.map(m=>[esc(m.model),esc(m.version),esc(m.provider),tags(m.files),m.reference_count]));
function renderPrompts(){
  const q=document.getElementById('qp').value.toLowerCase();
  const rows=DATA.prompts.filter(p=>!q||(p.name+' '+p.file+' '+p.preview+' '+p.type).toLowerCase().includes(q));
  renderTable('tPrompts',['Prompt','Type','Location','Purpose','Agent','Est. Tokens','Preview'],
    rows.map(p=>[esc(p.name),esc(p.type),esc(p.file)+':'+p.line,esc(p.purpose),esc(p.agent),
      p.estimated_tokens,esc(p.preview)]),{descCols:[6]});
}
document.getElementById('qp').oninput = renderPrompts; renderPrompts();
renderTable('tKeys',['Variable','Provider','AI Related','Used In','Loaded From'],
  DATA.api_keys.map(k=>[esc(k.variable),esc(k.provider),flag(k.ai_related),tags(k.used_in),tags(k.loaded_from)]));
renderTable('tTokens',['Agent','File','Model','Est. Input','Est. Output','Avg / Request',
  'Max Tokens','Temperature','Top P','Context Window','Estimated Cost'],
  DATA.token_usage.per_agent.map(t=>[esc(t.agent),esc(t.file),esc(t.model),
    esc(t.estimated_input_tokens),flag(t.estimated_output_tokens),flag(t.average_request_tokens),
    flag(t.max_tokens_configured),flag(t.temperature),flag(t.top_p),flag(t.context_window),
    flag(t.estimated_cost)]));
const RU = DATA.token_usage.runtime_usage;
renderTable('tRuntime',['Metric','Value'],
  [['Current Tokens Used',flag(RU.current_tokens_used)],['Today\'s Tokens Used',flag(RU.todays_tokens_used)],
   ['Monthly Tokens Used',flag(RU.monthly_tokens_used)],['Remaining Tokens',flag(RU.remaining_tokens)],
   ['Percentage Used',flag(RU.percentage_used)],['Instrumentation Found',tags(RU.instrumentation_found)],
   ['Instrumented Files',tags(RU.instrumentation_files)]]);
renderTable('tReq',['Metric','Value'],
  Object.entries(DATA.request_usage).map(([k,v])=>[esc(k.replace(/_/g,' ').replace(/\b\w/g,c=>c.toUpperCase())),
    Array.isArray(v)?tags(v):flag(v)]));
renderTable('tSdks',['SDK / Framework','Files','Count'],
  DATA.ai_sdks.map(s=>[esc(s.sdk),tags(s.files),s.file_count]));
renderTable('tTools',['Tool','Files','Count'],
  DATA.tools.map(t=>[esc(t.tool),tags(t.files),t.file_count]));
renderTable('tWf',['Workflow Pattern','Files','Count'],
  DATA.workflows.map(w=>[esc(w.pattern),tags(w.files),w.file_count]));
renderTable('tDeps',['File','Depends On'],
  Object.entries(DATA.dependencies.file_dependency_graph).map(([k,v])=>[esc(k),tags(v)]));
renderTable('tADeps',['Agent','File','Providers','Models','SDKs','Tools'],
  DATA.dependencies.agent_dependency_graph.map(a=>[esc(a.agent),esc(a.file),tags(a.providers),
    tags(a.models),tags(a.sdks),tags(a.tools)]));
renderTable('tOrph',['File'],DATA.dependencies.orphan_files.map(f=>[esc(f)]));
renderTable('tDead',['Symbol','Defined In','Note'],
  DATA.dependencies.dead_code_candidates.map(d=>[esc(d.symbol),esc(d.defined_in),esc(d.note)]));

/* ---------- theme + export ---------- */
function toggleTheme(){const h=document.documentElement;
  const n=h.dataset.theme==='dark'?'light':'dark';h.dataset.theme=n;
  try{localStorage.setItem('audit-theme',n)}catch(e){}}
try{const s=localStorage.getItem('audit-theme'); if(s) document.documentElement.dataset.theme=s;}catch(e){}
function download(name, text, mime){
  const b=new Blob([text],{type:mime}); const u=URL.createObjectURL(b);
  const a=document.createElement('a'); a.href=u; a.download=name; a.click(); URL.revokeObjectURL(u);}
function exportJSON(){download('repository_summary.json', JSON.stringify(DATA,null,2),'application/json');}
function exportCSV(){
  const cols=['serial','rel_path','name','ext','language','category','purpose','description',
    'size_human','lines_code','is_used','is_referenced','is_ai_related','dependencies'];
  const esc2=v=>'"'+String(Array.isArray(v)?v.join('; '):v==null?'':v).replace(/"/g,'""')+'"';
  const csv=[cols.join(',')].concat(DATA.file_inventory.map(f=>cols.map(c=>esc2(f[c])).join(','))).join('\n');
  download('repository_summary.csv', csv, 'text/csv');}
</script>
</body>
</html>
"""


def write_html(report: Dict[str, Any], path: str) -> None:
    payload = json.dumps(report, default=str)
    payload = payload.replace("</", "<\\/")  # keep the inline <script> safe
    html = (HTML_TEMPLATE
            .replace("__REPO__", _html.escape(report["repository"]["repository_name"]))
            .replace("__ROOT__", _html.escape(report["repository"]["repository_root"]))
            .replace("__DATE__", _html.escape(report["repository"]["scan_date"]))
            .replace("__AGENTS__", str(report["ai_agents"]["total_ai_agents"]))
            .replace("__VERSION__", VERSION)
            .replace("__NA__", NA)
            .replace("__DATA__", payload))
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(html)


# =====================================================================================
# SECTION 10 -- CLI
# =====================================================================================


def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        prog="ai_repo_audit.py",
        description="Complete static AI audit of a repository "
                    "(agents, providers, models, prompts, tokens, requests, structure).")
    ap.add_argument("root", nargs="?", default=".", help="Repository root (default: .)")
    ap.add_argument("-o", "--output", default=None,
                    help="Output directory (default: <root>/ai_audit_report)")
    ap.add_argument("--max-file-mb", type=float, default=3.0,
                    help="Max bytes read per file, in MB (default: 3)")
    ap.add_argument("--follow-symlinks", action="store_true", help="Follow symlinks")
    ap.add_argument("--no-xlsx", action="store_true", help="Skip the Excel report")
    ap.add_argument("--serve", nargs="?", const=8000, type=int, metavar="PORT",
                    help="Serve the report folder over http://127.0.0.1:PORT "
                         "(default 8000) instead of relying on file:// or another app")
    ap.add_argument("--open", dest="open_browser", action="store_true",
                    help="Open the dashboard in your default browser when done")
    ap.add_argument("--quiet", action="store_true", help="Suppress progress output")
    ap.add_argument("--version", action="version", version=f"AI Repository Auditor {VERSION}")
    args = ap.parse_args(argv)

    root = os.path.abspath(args.root)
    if not os.path.isdir(root):
        print(f"ERROR: not a directory: {root}", file=sys.stderr)
        return 2
    out_dir = os.path.abspath(args.output) if args.output \
        else os.path.join(root, "ai_audit_report")
    os.makedirs(out_dir, exist_ok=True)

    auditor = RepositoryAuditor(root, max_file_mb=args.max_file_mb, quiet=args.quiet,
                                follow_symlinks=args.follow_symlinks)
    auditor.log(f"AI Repository Auditor v{VERSION}")
    auditor.log(f"Scanning: {root}")
    auditor.walk()
    auditor.log(f"  files: {len(auditor.files)}  dirs: {len(auditor.dirs)}")
    auditor.log("Analysing dependencies, agents and AI usage ...")
    report = auditor.build_report()

    md_p = os.path.join(out_dir, "repository_summary.md")
    html_p = os.path.join(out_dir, "repository_summary.html")
    json_p = os.path.join(out_dir, "repository_summary.json")
    csv_p = os.path.join(out_dir, "repository_summary.csv")
    xlsx_p = os.path.join(out_dir, "repository_summary.xlsx")

    write_markdown(report, md_p)
    write_json(report, json_p)
    write_csv(report, csv_p)
    write_html(report, html_p)
    xlsx_written = None if args.no_xlsx else write_xlsx(report, xlsx_p)

    S = report["statistics"]
    auditor.log("")
    auditor.log("=" * 72)
    auditor.log(f"  Repository      : {report['repository']['repository_name']}")
    auditor.log(f"  Directories     : {S['total_directories']}")
    auditor.log(f"  Files           : {S['total_files']}  "
                f"(source {S['total_source_files']}, config {S['total_configuration_files']}, "
                f"docs {S['total_documentation_files']}, tests {S['total_test_files']}, "
                f"infra {S['total_infrastructure_files']})")
    auditor.log(f"  AI related files: {S['total_ai_related_files']}")
    auditor.log(f"  AI agents       : {S['total_agents']}")
    auditor.log(f"  Providers       : {S['total_providers']}  "
                f"({', '.join(p['provider'] for p in report['ai_providers']) or 'none'})")
    auditor.log(f"  Models          : {S['total_models']}")
    auditor.log(f"  SDKs            : {S['total_sdks']}")
    auditor.log(f"  Prompts         : {S['total_prompts']}")
    auditor.log(f"  Unused/orphan   : {len(report['dependencies']['orphan_files'])}")
    auditor.log("=" * 72)
    auditor.log("Reports written to: " + out_dir)
    for p in (md_p, html_p, json_p, csv_p):
        auditor.log("  - " + os.path.basename(p))
    if xlsx_written:
        auditor.log("  - " + os.path.basename(xlsx_p))
    elif not args.no_xlsx:
        auditor.log("  - repository_summary.xlsx  SKIPPED (install openpyxl to enable)")

    # The dashboard is a self-contained local file: it needs no web server and is NOT
    # served by any application in the audited repository. Print the exact file:// URL
    # so it is never mistaken for an app route (a wrong route yields "Not Found").
    file_url = "file://" + html_p.replace(os.sep, "/")
    auditor.log("")
    auditor.log("Open the dashboard directly (no web server required):")
    auditor.log("  " + file_url)

    if args.open_browser and not args.serve:
        try:
            import webbrowser
            webbrowser.open(file_url)
        except Exception as exc:  # pragma: no cover - environment dependent
            auditor.log(f"  (could not launch browser automatically: {exc})")

    if args.serve:
        _serve(out_dir, int(args.serve), auditor, args.open_browser)
    return 0


def _serve(out_dir: str, port: int, auditor: "RepositoryAuditor",
           open_browser: bool = False) -> None:
    """Serve the report directory locally, with the dashboard as the index page.

    This exists so the HTML can be viewed over http:// in environments where file://
    is blocked. It serves ONLY the generated report folder and never the repository.
    """
    import functools
    import http.server
    import socketserver

    class Handler(http.server.SimpleHTTPRequestHandler):
        def do_GET(self):  # noqa: N802 - stdlib naming
            if self.path in ("/", "/index.html", ""):
                self.path = "/repository_summary.html"
            return super().do_GET()

        def log_message(self, fmt, *a):  # keep output clean
            if not auditor.quiet:
                print("  http: " + fmt % a, flush=True)

    handler = functools.partial(Handler, directory=out_dir)
    socketserver.TCPServer.allow_reuse_address = True
    for attempt in range(20):
        try:
            httpd = socketserver.TCPServer(("127.0.0.1", port + attempt), handler)
            break
        except OSError:
            continue
    else:
        print(f"ERROR: no free port near {port}", file=sys.stderr)
        return
    bound = httpd.server_address[1]
    url = f"http://127.0.0.1:{bound}/"
    auditor.log("")
    auditor.log(f"Serving reports at {url}  (Ctrl+C to stop)")
    auditor.log(f"  dashboard : {url}")
    auditor.log(f"  json      : {url}repository_summary.json")
    auditor.log(f"  csv       : {url}repository_summary.csv")
    if open_browser:
        try:
            import webbrowser
            webbrowser.open(url)
        except Exception:
            pass
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        auditor.log("\nStopped.")
    finally:
        httpd.server_close()


if __name__ == "__main__":
    raise SystemExit(main())
